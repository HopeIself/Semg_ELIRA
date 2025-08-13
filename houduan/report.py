from fpdf import FPDF
import os
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook
from PIL import Image
import warnings
warnings.filterwarnings("ignore", message="cmap value too big/small*")


# 当前文件路径
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_REGULAR = os.path.join(CURRENT_DIR, "simsun.ttf")
FONT_BOLD = os.path.join(CURRENT_DIR, "SimSunBold.ttf")
CURRENT_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
CHINESE_NUM = {1:"一", 2:"二", 3:"三", 4:"四", 5:"五", 6:"六", 7:"七", 8:"八", 9:"九", 10:"十"}

action_map = {
    '握拳与打开手掌': '1',
    '手掌旋转': '2',
    '腕屈曲': '3',
    '腕伸展': '4',
    '手心向自己，手掌向内侧旋转': '5',
    '手心向自己，手掌向外侧旋转': '6',
    '压手': '7'
}

reaction_map = {
    "1": "握拳与打开手掌",
    "2": "手掌旋转",
    "3": "腕屈曲",
    "4": "腕伸展",
    "5": "手心向自己，手掌向内侧旋转",
    "6": "手心向自己，手掌向外侧旋转",
    "7": "压手"
}

def proceed_excel_sheets(excel_path):
    try:
        excel_file = pd.ExcelFile(excel_path)
        sheet_names = excel_file.sheet_names
        new_wb = load_workbook(excel_path)

        for sheet_name in sheet_names:
            if sheet_name in action_map:
                new_wb[sheet_name].title = action_map[sheet_name]
        new_wb.save(excel_path)
    except Exception as e:
        print(f"处理 Excel 文件时发生错误: {e}")

def add_std_emgplot(excel_path,output_path,num):
    col_names = ['肱三头肌', '前臂二', '前臂七']
    df = pd.read_excel(excel_path, usecols=[0, 1, 2], nrows=200, header=None, skiprows=1)
    plt.rcParams['font.sans-serif'] = ['SimSun']
    plt.rcParams['axes.unicode_minus'] = False
    plt.figure(figsize=(17, 5))
    
    for i in range(3):
        y = df.iloc[:, i]
        x = y.idxmax()
        y_max = y.max()
        plt.plot(y, label=col_names[i])
        plt.annotate(f"{y_max:.1f}", xy=(x, y_max), xytext=(x + 10, y_max + 10),
                     arrowprops=dict(arrowstyle='-', color='black'), fontsize=12, color='black')

    plt.title(f"标准动作{num}")
    ax = plt.gca()
    ax.set_xticklabels([])
    ax.tick_params(axis='both', which='both', length=0)
    plt.xlabel("时间（每格1s）")
    plt.ylabel("电位（μV）")
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    plot_path = os.path.join(output_path, f"标准动作{num}.png")
    plt.savefig(plot_path, dpi=300)
    plt.close()

def add_emgplot(excel_path,output_path):
    excel_file = pd.ExcelFile(excel_path)
    sheet_names = excel_file.sheet_names
    col_names = ['肱三头肌', '前臂二', '前臂七']
    
    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0, 1, 2], nrows=200, header=None, skiprows=1)
            plt.rcParams['font.sans-serif'] = ['SimSun']
            plt.rcParams['axes.unicode_minus'] = False
            plt.figure(figsize=(17, 5))
            
            for i in range(3):
                y = df.iloc[:, i]
                x = y.idxmax()
                y_max = y.max()
                plt.plot(y, label=col_names[i])
                plt.annotate(f"{y_max:.1f}", xy=(x, y_max), xytext=(x + 10, y_max + 10),
                             arrowprops=dict(arrowstyle='-', color='black'), fontsize=12, color='black')

            plt.title(f"{reaction_map[sheet_name]}")
            ax = plt.gca()
            ax.set_xticklabels([])
            ax.tick_params(axis='both', which='both', length=0)
            plt.xlabel("时间（每格1s）")
            plt.ylabel("电位（μV）")
            plt.grid(True)
            plt.legend()
            plot_path = os.path.join(output_path, f"{sheet_name}.png")
            plt.tight_layout()
            plt.savefig(plot_path, dpi=300)
            plt.close()
        except Exception as e:
            print(f"处理工作表 {sheet_name} 时发生错误: {e}")

def generate_report(data, image_map,muscle_map, output_path):
    pdf = EMGReport()
    pdf.add_page()
    pdf.add_basic_info(data['patient_info'])
    pdf.add_action_assessments(data['action_assessments'], image_map=image_map)
    pdf.add_muscle_assessments(data['std_training_data'], image_map=muscle_map)
    pdf.add_overall_assessment(data['overall_assessment'])
    pdf.add_model_info(data['model_info'])
    pdf.output(output_path)
    print(f"报告已保存至 {output_path}")

class EMGReport(FPDF):
    def __init__(self):
        super().__init__()
        if not os.path.exists(FONT_REGULAR):
            raise FileNotFoundError(f"未找到字体文件：{FONT_REGULAR}")
        if not os.path.exists(FONT_BOLD):
            raise FileNotFoundError(f"未找到字体文件：{FONT_BOLD}")
        self.add_font("SimSun", "", FONT_REGULAR, uni=True)
        self.add_font("SimSun", "B", FONT_BOLD, uni=True)
        self.set_auto_page_break(auto=True, margin=15)
        self.set_font("SimSun", size=12)

    def get_image_size(self, image_path):
        with Image.open(image_path) as img:
            return img.size

    def header(self):
        self.set_font("SimSun", 'B', 16)
        self.cell(0, 10, "肌电康复评估报告", ln=True, align="C")
        self.ln(3)

    def footer(self):
        self.set_y(-15)
        self.set_font("SimSun", '', 9)
        self.cell(0, 10, f'第 {self.page_no()} 页', align='C')

    def add_basic_info(self, info):
        self.set_font("SimSun", '', 12)
        self.cell(0, 5, f"报告时间：{CURRENT_DATE}", ln=True, align="L")
        self.cell(0, 10, f"姓名：{info['patient_id']}   性别：{info['gender']}   年龄：{info['age']}   主诉部位：{info['main_complaint']}", ln=True, align="L")
        self.ln(2)

    def add_action_assessments(self, assessments, image_map):
        self.set_font("SimSun", 'B', 13)
        self.cell(0, 10, "动作评估", ln=True, align="L", border='T')
        self.ln(3)

        for i, action in enumerate(assessments, 1):
            self.set_font("SimSun", 'B', 10)
            num = CHINESE_NUM.get(i, str(i))
            self.multi_cell(0, 8, f"动作{num}：{action['action_name']}")
            start_x, start_y = self.get_x(), self.get_y()
            image_height = 0
            info_text_height = 0

            if image_map and action['action_id'] in image_map:
                image_path = image_map[action['action_id']]
                if os.path.exists(image_path):
                    image_width, image_height_raw = self.get_image_size(image_path)
                    scale_factor = 140 / image_width
                    image_width = 140
                    image_height = image_height_raw * scale_factor
                    self.image(image_path, x=start_x, y=start_y, w=image_width)
                    info_x = start_x + image_width + 10
                    info_y = start_y
                    self.set_xy(info_x, info_y)
                    self.set_font("SimSun", '', 8)
                    line_height = 6
                    for metric in [f"MF: {action['MF']}", f"MNF: {action['MNF']}", f"RMS: {action['RMS']}", f"完成比例：{action['percent']}%"]:
                        self.set_xy(info_x, info_y + info_text_height)
                        self.cell(0, line_height, metric, ln=False)
                        info_text_height += line_height

            block_height = max(image_height, info_text_height)
            if block_height == 0:
                block_height = 10
            self.set_xy(start_x, start_y + block_height + 2)
            self.set_font("SimSun", '', 8)
            self.multi_cell(0, 8, f"评估：{action['evaluation']}")
            self.ln(4)
    
    def add_muscle_assessments(self, training_data, image_map): 
        self.set_font("SimSun", 'B', 13)
        self.cell(0, 10, "肌肉评估", ln=True, align="L", border='T')
        self.ln(3)

        for i, training in enumerate(training_data, 1):
            self.set_font("SimSun", 'B', 10)
            num = CHINESE_NUM.get(i, str(i))

            line_height = 6
            estimated_image_height = 0
            estimated_text_height = 6 * line_height

            if image_map and i in image_map:
                image_path = image_map[i]
                if os.path.exists(image_path):
                    image_width, image_height_raw = self.get_image_size(image_path)
                    scale_factor = 140 / image_width
                    estimated_image_height = image_height_raw * scale_factor

            estimated_block_height = max(estimated_image_height, estimated_text_height)
            if estimated_block_height == 0:
                estimated_block_height = 10

            if self.get_y() + estimated_block_height + 20 > self.page_break_trigger:
                self.add_page()
                self.ln(3)
            self.set_font("SimSun", 'B', 10)
            self.multi_cell(0, 8, f"标准动作{num}：")
            start_x, start_y = self.get_x(), self.get_y()
            image_height = 0
            info_text_height = 0

            if image_map and i in image_map:
                image_path = image_map[i]
                if os.path.exists(image_path):
                    image_width, image_height_raw = self.get_image_size(image_path)
                    scale_factor = 140 / image_width
                    image_width = 140
                    image_height = image_height_raw * scale_factor
                    self.image(image_path, x=start_x, y=start_y, w=image_width)
                    info_x = start_x + image_width + 10
                    info_y = start_y
                    self.set_xy(info_x, info_y)
                    self.set_font("SimSun", '', 8)
            
                    # 获取训练数据
                    values = [
                        training['yesterday_progress'],
                        training['three_days_ago_progress'],
                        training['seven_days_ago_progress']
                    ]
                    time_labels = ["较昨日", "较三天前", "较七天前"]
            
                    # 如果数据大于 2，则写入 "暂无报告参考"
                    units = []
                    for value, label in zip(values, time_labels):
                        if value > 5:
                            unit = f"{label} 暂无报告参考"
                        elif value > 1:
                            unit = f"{label} ↑ {(value - 1) * 100:.1f}%"
                        else:
                            unit = f"{label} ↓ {(1 - value) * 100:.1f}%"
                        units.append(unit)
            
                    # 设置并显示其他指标
                    metrics = [f"MF: {training['MF']}", f"MNF: {training['MNF']}", f"RMS: {training['RMS']}"]
                    for metric in metrics:
                        self.set_xy(info_x, info_y + info_text_height)
                        self.cell(0, line_height, metric, ln=False)
                        info_text_height += line_height
            
                    # 显示单位（包括“暂无报告参考”）
                    for unit in units:
                        self.set_xy(info_x, info_y + info_text_height)
                        self.cell(0, line_height, unit, ln=False)
                        info_text_height += line_height

            block_height = max(image_height, info_text_height)
            if block_height == 0:
                block_height = 10
            self.set_xy(start_x, start_y + block_height + 2)
            self.ln(4)




    
    def add_overall_assessment(self, overall):
        self.set_font("SimSun", 'B', 13)
        self.cell(0, 10, '总体评估与康复建议', ln=True, align="L", border='T')
        self.ln(2)
        self.set_font("SimSun", 'B', 10)
        self.cell(0, 8, f"当前阶段：{overall['rehab_stage']}", ln=True)
        self.set_font("SimSun", '', 9)
        self.multi_cell(0, 8, f"总结：{overall['summary']}")
        self.ln(2)
        self.set_font("SimSun", 'B', 10)
        self.cell(0, 8, '建议方案：', ln=True)
        self.set_font("SimSun", '', 9)
        for item in overall['clinical_recommendation']:
            self.multi_cell(0, 8, f"·{item}")
        self.cell(0, 5, '', ln=True, align="L", border='B')

    def add_model_info(self, model_info):
        self.set_font("SimSun", '', 9)
        self.cell(50, 8, f"分析模型：{model_info['inference_model']}", ln=False, align="L")
        self.multi_cell(0, 8, f"{model_info['notes']}", align="R")


def write_ai_comment_to_excel(user_name, ai_comments):
    excel_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "User.xlsx")
    
    if not os.path.exists(excel_file_path):
        print(f"未找到用户数据文件: {excel_file_path}")
        return

    try:
        df = pd.read_excel(excel_file_path)
        
        if "name" not in df.columns or "ai_com" not in df.columns:
            print("Excel中缺少 'name' 或 'ai_com' 列")
            return

        # 仅提取建议4（注意索引为3）
        if len(ai_comments) >= 4:
            advice_4 = ai_comments[3]
        else:
            advice_4 = "暂无建议4"

        # 更新对应 name 行的 ai_com 列
        df.loc[df["name"] == user_name, "ai_com"] = advice_4

        # 保存
        df.to_excel(excel_file_path, index=False)
        print(f"建议4已成功写入 {excel_file_path}")

    except Exception as e:
        print(f"写入建议4失败：{e}")

import os
from datetime import datetime

def generate_all_from_excel(excel_path, std_paths, report_data, user_name):
    # 获取当前文件夹路径
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建目标报告存储路径
    user_report_path = os.path.join(current_dir, 'user_files', user_name, 'Reports')
    
    # 确保目标目录存在
    os.makedirs(user_report_path, exist_ok=True)

    # 调用其他处理函数
    proceed_excel_sheets(excel_path)
    for idx, std_path in enumerate(std_paths, 1):
        add_std_emgplot(std_path, current_dir, idx)
    
    add_emgplot(excel_path, current_dir)

    # 创建图像映射
    image_map = {
        a['action_id']: os.path.join(current_dir, f"{a['action_id']}.png")
        for a in report_data['action_assessments']
    }

    # 创建标准动作图像映射
    muscle_map = {
        idx + 1: os.path.join(current_dir, f"标准动作{idx + 1}.png") for idx in range(len(std_paths))
    }

    # 生成报告文件名
    output_filename = f"{datetime.now().strftime('%Y%m%d%H%M')}_肌电评估报告.pdf"

    # 完整报告文件路径
    output_path = os.path.join(user_report_path, output_filename)

    # 生成报告
    generate_report(report_data, image_map, muscle_map, output_path)
    
    write_ai_comment_to_excel(user_name, report_data['overall_assessment']['clinical_recommendation'])
    

    print(f"肌电报告生成成功：{output_path}")
