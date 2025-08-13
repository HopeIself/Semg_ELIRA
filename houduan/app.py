# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, make_response, Response, session, current_app, abort , send_from_directory # 导入 make_response
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
import os
from cozepy import JWTOAuthApp
import json
from cryptography.hazmat.primitives import serialization
import logging
import time
import openpyxl
import uuid
import socket
import pandas as pd
from werkzeug.utils import secure_filename
import threading
import matplotlib
import requests
import logging
from services.ai_service import select_ai_model_api
from services.emg_service import run_initial_assessment, process_data_and_train_model, run_prediction_session
import numpy as np
import queue  # 新增
from utils.global_queue import UDP_DATA_QUEUE  # 修改：从全局队列模块导入
from utils.config import ACTION_MAP  # 保持原有导
from services.ai_service import AI_MODELS
from models.deepseek_model import DeepSeekSession
from models.tongyi_model import TongyiSession
from models.volcano_model import VolcanoSession
from openpyxl import Workbook
import datetime
import glob
from cozepy import COZE_CN_BASE_URL
import openai
from report import generate_all_from_excel
import AccessToken
from AccessToken import AccessToken as RTCAccessToken

# 火山引擎配置 - 替换为你的实际信息
RTC_APP_ID = "688c128b3f54ba01771e6b2f"  # 从火山引擎控制台获取的AppID
# 火山引擎配置 - 替换为你的实际信息
RTC_APP_KEY = "952e01e83cce40dc8351aedd38fdd106"       # 从火山引擎控制台获取的AppKey
# ==== 配置区域 ====
client_id = "1136790881240"
public_key_id = "2bZn7ZDMslOIRwomJbQOkLMytzOBQ5V-pefCFWDCF-8"
api_base = COZE_CN_BASE_URL  # or "https://api.coze.com"
private_key_path = "private_key.pem"
token_url = "https://api.coze.cn/api/permission/oauth2/token"


global CURRENT_AI_SESSION  # Make sure this is accessible globally
# 顶部初始化（放在 app.py 顶部或合适位置）
CURRENT_AI_SESSION: dict[int, tuple[object, str]] = {}

CURRENT_DIR = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


# ── 日志配置 ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,                              # 改成 DEBUG 会更啰嗦
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

USER_FILE = 'User.xlsx'
STD_DIR = 'user_files'  # 存储用户文件的根目录


# 初始化 Flask 应用
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})  # 启用 CORS
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
jwt = JWTManager(app)

excel_file = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'user_data.xlsx')

# def monitor_udp_queue():
#     """实时打印队列信息"""
#     import time
#     while True:
#         queue_size = UDP_DATA_QUEUE.qsize()
#         print(f"[UDP数据队列] 当前队列长度: {queue_size}")

#         # 可选：打印队列前几个元素，注意队列是线程安全的，但要避免破坏数据
#         try:
#             temp_list = list(UDP_DATA_QUEUE.queue)[:5]
#             print(f"[UDP数据队列] 前5个数据: {temp_list}")
#         except Exception as e:
#             print(f"[UDP数据队列] 读取内容失败: {e}")

#         time.sleep(1)  # 每秒打印一次

#依据原始数据计算特征值
def calculate_emg_features(gesture_data):
    # gesture_data: shape (N, 3)
    if gesture_data.ndim == 1:
        gesture_data = gesture_data.reshape(-1, 1)
    # 对每个通道分别计算
    rms_list, mnf_list, mf_list = [], [], []
    for ch in range(gesture_data.shape[1]):
        x = gesture_data[:, ch]
        rms = np.sqrt(np.mean(np.square(x)))
        # FFT
        fft_result = np.fft.fft(x)
        freq_bins = np.fft.fftfreq(len(x))
        pos_mask = freq_bins >= 0
        freq_bins = freq_bins[pos_mask]
        fft_result = fft_result[pos_mask]
        # MNF
        power = np.abs(fft_result) ** 2
        mnf = np.sum(freq_bins * power) / (np.sum(power) + 1e-8)
        # MF
        abs_fft = np.abs(fft_result)
        mf = np.sum(freq_bins * abs_fft) / (np.sum(abs_fft) + 1e-8)
        rms_list.append(rms)
        mnf_list.append(mnf)
        mf_list.append(mf)
    # 返回均值，保证是float
    return float(np.mean(rms_list)), float(np.mean(mnf_list)), float(np.mean(mf_list))
    
# 确保 Excel 文件存在
def create_excel_file():
    if not os.path.exists(excel_file):
        # 如果文件不存在，创建文件并添加列标题
        df = pd.DataFrame(columns=["id", "email", "password"])
        df.to_excel(excel_file, index=False)
        

current_sse_generator = None
sse_lock = threading.Lock()  # 锁，用于确保每次只处理一个连接


def process_udp_data(data):
    """
    处理接收到的 UDP 数据
    """
    try:
        obj = json.loads(data.decode())  # 尝试解析数据
        # 如果收到的是list，直接入队每个元素
        if isinstance(obj, list):
            for emg_values in obj:
                try:
                    UDP_DATA_QUEUE.put_nowait(emg_values)  # 尝试将数据放入队列
                except queue.Full:
                    # 如果队列满了，丢弃最旧的数据
                    try:
                        UDP_DATA_QUEUE.get_nowait()
                    except queue.Empty:
                        pass
                    UDP_DATA_QUEUE.put_nowait(emg_values)
    except Exception as e:
        print(f"UDP 数据解析失败: {e}")


#用于测试用户是否佩戴好        
@app.route('/api/real-time-feedback', methods=['GET'])
def real_time_feedback():
    """
    每秒两次向前端反馈实时肌电信号的绝对值的最大值，如果超过100则反馈"好"，否则反馈"不好"。
    使用SSE实时推送数据。
    """

    def event_stream():
        import time
        import numpy as np

        while True:
            try:
                # 从队列获取最新的肌电数据
                emg_data = UDP_DATA_QUEUE.get(timeout=2)

                # 确保emg_data是一个数组
                if isinstance(emg_data, list):
                    emg_data = np.array(emg_data)  # 如果是列表，将其转为numpy数组

                # 计算每个通道的绝对值
                emg_abs = np.abs(emg_data)  # 获取每个通道的绝对值

                # 计算绝对值中的最大值
                max_emg_value = np.max(emg_abs)  # 取三个通道中最大绝对值

                # 根据最大绝对值判断反馈内容
                feedback = "好" if max_emg_value > 100 else "不好"

                # 发送数据给前端
                yield f"data: {json.dumps({'emg': emg_data.tolist(), 'max_emg_value': max_emg_value, 'feedback': feedback})}\n\n"
                
                # 等待0.5秒，确保每秒反馈两次
                time.sleep(0.25)
            except queue.Empty:
                yield "data: {}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'error': f'发生错误: {str(e)}'})}\n\n"

    return Response(event_stream(), mimetype='text/event-stream')


#用于进行标准评估与肌肉评估
@app.route('/api/initial_assessment', methods=['GET'])
def initial_assessment():
    user_id = request.args.get('id')
    code = request.args.get('code')
    logger.debug(f"Received request with user_id: {user_id}, code: {code}")

    if not user_id or not code:
        logger.error("缺少用户ID或code")
        return jsonify({"error": "缺少用户ID或code"}), 400

    user_name = get_user_name(user_id)
    if not user_name:
        logger.error(f"用户ID {user_id} 无效")
        return jsonify({"error": "用户ID无效"}), 404

    folder_name = "Initial_assessment" if code == '1' else "Mulscle_assessment" if code == '2' else None
    if folder_name is None:
        return jsonify({"error": "无效的code"}), 400

    user_folder = os.path.join('user_files', user_name, folder_name)
    os.makedirs(user_folder, exist_ok=True)

    def event_stream():
        collected_data = {f"{i}_{metric}": None for i in (1, 2) for metric in ["RMS", "MNF", "MF"]}
        gesture_data_1, gesture_data_2 = [], []
        user_data = get_user_data(user_id)
        if user_data is None:
            logger.error(f"无法获取用户 {user_id} 的数据")
            return

        # 倒计时
        for i in range(3, 0, -1):
            yield f"data: {json.dumps({'countdown': i})}\n\n"
            time.sleep(1)
        
        # 第一次采集
        for sec in range(10, 0, -1):
            # 清理数据队列，确保读取新的数据
            while not UDP_DATA_QUEUE.empty():
                try:
                    UDP_DATA_QUEUE.get_nowait()
                except queue.Empty:
                    break
            
            emg_one_sec = []
            start = time.time()
            
            # 收集一秒钟的数据，确保每秒数据完整
            while time.time() - start < 1:
                try:
                    emg_data = UDP_DATA_QUEUE.get(timeout=0.2)
                    if isinstance(emg_data, list):
                        emg_one_sec.append(np.array(emg_data))
                except queue.Empty:
                    break
            
            if emg_one_sec:
                emg_all = np.concatenate(emg_one_sec)
                rms, mnf, mf = calculate_emg_features(emg_all)
                
                if collected_data["1_RMS"] is None:
                    collected_data["1_RMS"] = rms
                    collected_data["1_MNF"] = mnf
                    collected_data["1_MF"] = mf
                
                gesture_data_1.extend(emg_one_sec)
        
                # 只发送每秒内的第一个数据块（例如，`emg_one_sec[0]`）
                yield f"data: {json.dumps({'emg': emg_one_sec[0].tolist(), 'countdown': sec})}\n\n"
            else:
                # 如果 emg_one_sec 为空，发送空数据
                yield f"data: {json.dumps({'emg': [], 'countdown': sec})}\n\n"

        
        
        time.sleep(1)
        # 中间倒计时
        for i in range(5, 0, -1):
            yield f"data: {json.dumps({'countdown': i})}\n\n"
            time.sleep(1)

        # 第二次采集
        for sec in range(10, 0, -1):
            # 清理数据队列，确保读取新的数据
            while not UDP_DATA_QUEUE.empty():
                try:
                    UDP_DATA_QUEUE.get_nowait()
                except queue.Empty:
                    break
            
            emg_one_sec = []
            start = time.time()
            
            # 收集一秒钟的数据，确保每秒数据完整
            while time.time() - start < 1:
                try:
                    emg_data = UDP_DATA_QUEUE.get(timeout=0.2)
                    if isinstance(emg_data, list):
                        emg_one_sec.append(np.array(emg_data))
                except queue.Empty:
                    break
            
            if emg_one_sec:
                emg_all = np.concatenate(emg_one_sec)
                rms, mnf, mf = calculate_emg_features(emg_all)
                
                if collected_data["2_RMS"] is None:
                    collected_data["2_RMS"] = rms
                    collected_data["2_MNF"] = mnf
                    collected_data["2_MF"] = mf
                
                gesture_data_2.extend(emg_one_sec)
        
                # 只发送每秒内的第一个数据块（例如，`emg_one_sec[0]`）
                yield f"data: {json.dumps({'emg': emg_one_sec[0].tolist(), 'countdown': sec})}\n\n"
            else:
                # 如果 emg_one_sec 为空，发送空数据
                yield f"data: {json.dumps({'emg': [], 'countdown': sec})}\n\n"

        
        time.sleep(1)
        
        yield f"data: {json.dumps({'done': 5})}\n\n"

        # 保存数据
        save_data_to_excel(user_folder, collected_data, user_id, user_data["ai_plan"])

        if code == '2':
            training_data_folder = os.path.join(user_folder, "Training_data")
            os.makedirs(training_data_folder, exist_ok=True)

            def save_labeled_emg_block(data_list, block_index):
                df = pd.DataFrame([x.tolist() for x in data_list])
                df.columns = [1, 2, 3]  # 第一行为1,2,3
                now = datetime.datetime.now()
                file_path = os.path.join(training_data_folder, f"{now.year}年{now.month:02d}月{now.day:02d}日_Training-data_{block_index}.xlsx")
                df.to_excel(file_path, index=False, header=True)
                logger.info(f"Saved block {block_index} to {file_path}")

            save_labeled_emg_block(gesture_data_1, 1)
            save_labeled_emg_block(gesture_data_2, 2)
            yield f"data: {json.dumps({'done': 2})}\n\n"

        if code == '1':
            send_user_info(user_id, user_data, collected_data)
            
            yield f"data: {json.dumps({'done': 1})}\n\n"


    return Response(event_stream(), mimetype='text/event-stream')


def get_user_name(user_id):
    """从User.xlsx中查找用户的姓名"""
    try:
        df = pd.read_excel('User.xlsx')
        logger.debug(f"Data from User.xlsx: {df.head()}")  # 调试：打印前几行数据

        user_row = df[df['userId'] == user_id]
        if not user_row.empty:
            user_name = user_row.iloc[0]['name']  # 假设用户姓名列是 'name'
            logger.debug(f"Found user name for {user_id}: {user_name}")
            return user_name
        else:
            logger.warning(f"User with ID {user_id} not found.")
            return None
    except Exception as e:
        logger.error(f"读取User.xlsx失败: {e}")
        return None


def get_user_data(user_id):
    """从User.xlsx中获取用户的详细信息"""
    try:
        df = pd.read_excel('User.xlsx')
        user_row = df[df['userId'] == user_id]
        if not user_row.empty:
            user_data = user_row.iloc[0]
            return {
                "1_RMS_std": user_data["1_RMS_std"],
                "1_MNF_std": user_data["1_MNF_std"],
                "1_MF_std": user_data["1_MF_std"],            
                "2_RMS_std": user_data["2_RMS_std"],
                "2_MNF_std": user_data["2_MNF_std"],
                "2_MF_std": user_data["2_MF_std"],
                "ache": user_data["ache"],
                "ai_plan": user_data["ai_plan"],
                "doc_com": user_data["doc_com"],
                "percent1" : user_data["percent1"],
                "percent2" : user_data["percent2"],
                "percent3" : user_data["percent3"],
                "ai_com"  : user_data["ai_com"],
                "doc_judge": user_data["doc_judge"],
                "age"      : user_data["age"],
                "gender"   : user_data["gender"]
            }
        return None
    except Exception as e:
        logger.error(f"读取User.xlsx失败: {e}")
        return None


def send_user_info(user_id, user_data, collected_data):
    """发送用户信息和采集数据到AI，并接收AI的反馈"""
    global CURRENT_AI_SESSION  # Make sure this is accessible globally
    # 整理嵌套结构的数据
    payload = {
        "user": {
            "initial_amount": [
                {
                    "MF": collected_data["1_MF"],
                    "MNF": collected_data["1_MNF"],
                    "RMS": collected_data["1_RMS"]
                },
                {
                    "MF": collected_data["2_MF"],
                    "MNF": collected_data["2_MNF"],
                    "RMS": collected_data["2_RMS"]
                }
            ],
            "normal_amount": [
                {
                    "MF": user_data["1_MF_std"],
                    "MNF": user_data["1_MNF_std"],
                    "RMS": user_data["1_RMS_std"]
                },
                {
                    "MF": user_data["2_MF_std"],
                    "MNF": user_data["2_MNF_std"],
                    "RMS": user_data["2_RMS_std"]
                }
            ],
            "percent1" : user_data["percent1"],
            "percent2" : user_data["percent2"],
            "percent3" : user_data["percent3"],
            "ache": user_data["ache"],
            "used_plan": user_data["ai_plan"],
            "doc_com": user_data["doc_com"],
            "ai_com": user_data["ai_com"],
            "doc_judge": user_data["doc_judge"],
            "age"      : user_data["age"],
            "gender"   : user_data["gender"]
        }
    }

    # === ✅ 关键改动：确保 JSON 可序列化 ===
    payload_cleaned = convert_numpy_types(payload)
    logger.info(f"发送给AI的数据: {json.dumps(payload_cleaned, ensure_ascii=False)}")

    # 发送数据给AI并接收反馈
    feedback = None
    if 'CURRENT_AI_SESSION' in globals() and CURRENT_AI_SESSION:
        feedback = CURRENT_AI_SESSION.send_input(payload_cleaned)

    if feedback:
        logger.info(f"Received feedback from AI: {feedback}")

        # 提取 AI 返回的计划数据
        if 'actions' in feedback and 'message' in feedback:
            ai_plan = {
                "type": "plan",
                "actions": feedback['actions'],
                "message": feedback['message']
            }
            logger.info(f"Received valid ai_plan: {ai_plan}")
            update_ai_plan_in_excel(user_id, ai_plan)
        else:
            logger.warning(f"AI did not return a valid ai_plan. Received feedback: {feedback}")

    return feedback


def convert_numpy_types(obj):
    """递归转换 numpy 类型为原生类型"""
    if isinstance(obj, dict):
        return {k: convert_numpy_types(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(i) for i in obj]
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    else:
        return obj


def update_ai_plan_in_excel(user_id, ai_plan):
    """更新 User.xlsx 中的 ai_plan 列"""
    try:
        df = pd.read_excel('User.xlsx')

        # 确保 user_id 和 userId 列的数据类型一致
        if isinstance(user_id, str):
            df['userId'] = df['userId'].astype(str)
        elif isinstance(user_id, int):
            df['userId'] = df['userId'].astype(int)

        # 查找对应用户的行
        user_row = df[df['userId'] == user_id]
        
        if not user_row.empty:
            logger.info(f"当前用户 {user_id} 的 ai_plan: {user_row['ai_plan'].iloc[0]}")
            
            # 确保 ai_plan 有效
            if ai_plan is not None and ai_plan != "":
                # 更新 ai_plan 列
                ai_plan_cleaned = convert_numpy_types(ai_plan)
                df.loc[df['userId'] == user_id, 'ai_plan'] = json.dumps(ai_plan_cleaned, ensure_ascii=False)

                logger.info(f"更新后的用户 {user_id} 的 ai_plan: {ai_plan}")
                
                # 保存更新后的 DataFrame 到 Excel 文件
                try:
                    df.to_excel('User.xlsx', index=False)
                    logger.info(f"成功保存 ai_plan 到 User.xlsx")
                except Exception as e:
                    logger.error(f"保存文件时出错: {e}")
            else:
                logger.warning(f"ai_plan 无效，无法更新：{ai_plan}")
        else:
            logger.warning(f"在 User.xlsx 中找不到 ID 为 {user_id} 的用户")
    
    except Exception as e:
        logger.error(f"更新 ai_plan 时出错: {e}")


def save_data_to_excel(user_folder, collected_data, user_id, ai_plan):
    """保存数据到STD.xlsx和日期文件，并更新 ai_plan"""
    try:
        std_file = os.path.join(user_folder, 'STD.xlsx')
        if os.path.exists(std_file):
            df = pd.read_excel(std_file)
            if df.empty:
                logger.info("STD.xlsx is empty. Creating new file for today.")
                create_and_save_data(df, std_file, collected_data)
            else:
                date_file_name = f"{datetime.datetime.now().strftime('%Y年%m月%02d日')}.xlsx"
                date_file_path = os.path.join(user_folder, date_file_name)
                logger.info(f"STD.xlsx has data. Creating new file: {date_file_path}")
                create_and_save_data(None, date_file_path, collected_data)
        else:
            logger.info(f"{std_file} not found, creating STD.xlsx and saving data.")
            create_and_save_data(None, std_file, collected_data)

    except Exception as e:
        logger.error(f"Error saving data to Excel: {e}")


def create_and_save_data(df, file_path, collected_data):
    """创建并保存数据到指定文件"""
    new_data = pd.DataFrame([{
        "1_RMS": collected_data["1_RMS"],
        "1_MNF": collected_data["1_MNF"],
        "1_MF": collected_data["1_MF"],
        "2_RMS": collected_data["2_RMS"],
        "2_MNF": collected_data["2_MNF"],
        "2_MF": collected_data["2_MF"]
    }])

    if df is not None:  # 如果已有数据，追加
        df = pd.concat([df, new_data], ignore_index=True)
    else:  # 如果没有数据，则创建新文件
        df = new_data

    df.to_excel(file_path, index=False)
    logger.info(f"Data saved to {file_path}")

def udp_server():
    """
    UDP 服务器函数，接收数据并处理
    """
    UDP_IP = "0.0.0.0"  # 监听所有网络接口
    UDP_PORT = 6000  # 目标端口，确保与前端一致
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)  # 允许端口重用
    sock.bind((UDP_IP, UDP_PORT))

    print(f"Listening for UDP packets on {UDP_IP}:{UDP_PORT}...")

    while True:
        data, addr = sock.recvfrom(4096)  # 接收最多 4096 字节的数据
        threading.Thread(target=process_udp_data, args=(data,)).start()  # 使用线程处理每个数据包


def start_udp_server():
    """
    启动 UDP 服务器的线程
    """
    udp_thread = threading.Thread(target=udp_server)
    udp_thread.daemon = True
    udp_thread.start()



# 你可以通过调用 start_udp_server 来启动 UDP 服务器
    # # 启动队列监控线程


# 在你的程序启动时调用 start_udp_server()



@app.route('/api/user/register', methods=['POST'])
def register_user():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    role = data.get('role')  # 接收角色

    if not email or not password or not role:
        return jsonify({"message": "邮箱、密码和角色不能为空"}), 400

    # 读取 Excel 文件，检查该邮箱是否已注册
    try:
        df = pd.read_excel(excel_file)
    except FileNotFoundError:
        return jsonify({"message": "文件未找到，请检查路径设置"}), 500

    # 查找是否已有该邮箱，且该邮箱的密码和角色都匹配
    existing_user = df[(df['email'] == email) & (df['password'] == password) & (df['role'] == role)]
    
    if not existing_user.empty:
        return jsonify({"message": "该邮箱、密码和角色已被注册"}), 400

    # 查找是否有重复的邮箱
    existing_email = df[df['email'] == email]
    if not existing_email.empty:
        return jsonify({"message": "该邮箱已被注册"}), 400

    # 生成用户唯一的 ID（可以使用 UUID）
    user_id = str(uuid.uuid4())  # 这里使用 UUID 生成唯一的用户 ID

    # 将用户信息写入 Excel 文件
    new_user = pd.DataFrame({"id": [user_id], "email": [email], "password": [password], "role": [role]})
    df = pd.concat([df, new_user], ignore_index=True)
    df.to_excel(excel_file, index=False)

    return jsonify({"message": "注册成功"}), 200


# 登录接口
@app.route('/api/user/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        # 响应 preflight 请求
        response = make_response()
        response.headers['Access-Control-Allow-Origin'] = '*'  # 允许所有来源
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'  # 允许的方法
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'  # 允许的请求头
        return response

    # 处理 POST 请求
    email = request.json.get('email', None)
    password = request.json.get('password', None)
    role = request.json.get('role', None)

    # 调试日志：打印接收到的 email 和 password
    print(f"Received email: {email}, password: {password}, role: {role}")

    if not email or not password or not role:
        return jsonify({"error": "邮箱、密码和角色不能为空"}), 400

    # 读取 Excel 文件，检查该邮箱是否已注册
    try:
        df = pd.read_excel(excel_file)
    except FileNotFoundError:
        return jsonify({"message": "文件未找到，请检查路径设置"}), 500

    # 查找用户
    user_row = df[df['email'] == email]
    
    if user_row.empty:
        print(f"Invalid credentials for email: {email}")  # 打印无效的凭据
        return jsonify({"error": "无效的邮箱或密码"}), 401
    
    # 获取用户信息
    user = user_row.iloc[0]  # 获取该用户的第一行数据

    # 验证密码
    if user['password'] != password:
        print(f"Invalid credentials for email: {email}")  # 打印无效的凭据
        return jsonify({"error": "无效的邮箱或密码"}), 401

    # 验证角色
    if user['role'] != role:
        print(f"Role mismatch for email: {email}, expected: {user['role']}, received: {role}")  # 打印角色不匹配
        return jsonify({"error": "角色不匹配"}), 403

    # 在成功登录时，返回包括 id 和 token 的数据
    return jsonify({
        "code": 1,
        "data": {
            "id": user['id'],  # 假设用户 ID 是 user['id']
        }
    }), 200



# 受保护的接口（验证 Token）
@app.route('/protected', methods=['GET'])
@jwt_required()
def protected():
    """需要 JWT 授权的受保护路由"""
    # 获取当前用户的身份信息
    current_user = get_jwt_identity()
    return jsonify(logged_in_as=current_user), 200

@app.route('/api/models', methods=['GET'])
def get_available_models():
    """获取可用的AI模型列表"""
    from services.ai_service import get_available_models
    return jsonify(get_available_models())

@app.route('/api/select-model', methods=['POST'])
def select_model():
    """选择AI模型"""
    global CURRENT_AI_SESSION

    data = request.json
    model_id = data.get('model_id')
    api_key = data.get('api_key')

    if not model_id:
        return jsonify({"error": "需要提供model_id参数"}), 400

    try:
        CURRENT_AI_SESSION = select_ai_model_api(model_id, api_key)
        return jsonify({"success": True, "message": f"已成功选择{model_id}模型"})
    except Exception as e:
        logger.error(f"选择模型失败: {str(e)}")
        return jsonify({"error": f"选择模型失败: {str(e)}"}), 500

@app.route('/api/get-training-plan', methods=['POST'])
def get_training_plan():
    """
    第一次访问，前端发送 id，后端返回训练计划中的动作。
    前端发送 id，后端去 User.xlsx 获取训练计划并返回相应动作。
    """
    global GESTURE_LIST

    GESTURE_LIST = []  # 清空手势列表

    logger.info("进入 get_training_plan()")
    logger.info(f"当前 session.keys(): {list(session.keys())}")

    # 获取 user_id
    user_id = None
    if request.is_json:
        user_id = request.json.get('id')

    logger.debug(f"接收到的用户ID: {user_id}")

    if not user_id:
        logger.warning("未提供 userId")
        return jsonify({"error": "缺少用户id"}), 400

    # 从 User.xlsx 读取 ai_plan
    user_xlsx = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'User.xlsx')
    logger.debug(f"检查 User.xlsx 文件路径: {user_xlsx}")
    
    if not os.path.exists(user_xlsx):
        logger.error("User.xlsx 文件不存在")
        return jsonify({"error": "User.xlsx 文件不存在"}), 500

    try:
        # 打开并读取 Excel 文件
        logger.debug("开始读取 Excel 文件")
        df = pd.read_excel(user_xlsx)
        logger.debug(f"读取到的用户数据：\n{df.head()}")  # 输出文件的前几行，检查数据

        idx = df.index[df['userId'] == user_id].tolist()
        if not idx:
            logger.warning(f"User.xlsx 未找到 id={user_id}")
            return jsonify({"error": "用户不存在"}), 404

        row_idx = idx[0]
        logger.debug(f"找到用户 {user_id} 对应的训练计划行索引: {row_idx}")

        ai_plan_str = df.at[row_idx, 'ai_plan']
        if not ai_plan_str or pd.isna(ai_plan_str):
            logger.warning(f"用户 {user_id} 没有 ai_plan")
            return jsonify({"error": "该用户没有训练计划，请先进行初始评估"}), 400

        try:
            ai_plan = json.loads(ai_plan_str)
            logger.debug(f"加载的训练计划: {ai_plan}")
        except Exception as e:
            logger.error(f"ai_plan 字段解析失败: {e}")
            return jsonify({"error": "训练计划格式错误"}), 500

        actions = ai_plan.get('actions', [])
        logger.debug(f"训练计划中的动作: {actions}")
        if not actions:
            logger.warning("训练计划中没有动作")
            return jsonify({"error": "训练计划中没有动作"}), 400


        # 为每个动作生成名称，并返回训练计划
        for action in actions:
            if 'name' not in action:
                # 根据 action_id 动态生成动作名称
                action['name'] = get_action_name_by_id(action['action_id'])
                logger.debug(f"为动作 {action['action_id']} 自动生成名称: {action['name']}")

        # 返回训练计划中的动作
        result = {
            "message": ai_plan.get("message", "训练计划信息未提供"),
            "actions": [{"name": action['name'], "time": action.get('time', 0)} for action in actions]
        }
        logger.debug(f"返回的训练计划结果: {result}")
        return jsonify(result)

    except Exception as e:
        logger.error(f"训练失败: {str(e)}")
        return jsonify({"error": f"训练失败: {str(e)}"}), 500


def get_action_name_by_id(action_id):
    """
    根据 action_id 动态生成动作名称。
    """
    action_names = {
        1: '握拳与打开手掌',
        2: '手掌旋转',
        3: '腕屈曲',
        4: '腕伸展',
        5: '手心向自己，手掌向内侧旋转',
        6: '手心向自己，手掌向外侧旋转',
        7: '压手'
    }
    return action_names.get(action_id, f"未知动作 {action_id}")


@app.route('/api/start-training-process')
def start_training_process():
    """
    SSE推送：训练/休息阶段、剩余秒数、每0.5秒实时肌电值
    增加详细调试日志，并修正推送给前端的秒数准确性
    并将特征值写入 user_files/{name}/training/std.xlsx
    每一次采集前清空队列
    修正采集数据点数异常波动问题
    """
    user_id = request.args.get('id')
    logger.info(f"[SSE] 进入start_training_process user_id={user_id}")

    def sse_stream():
        logger.info("[SSE] sse_stream启动")
        if not hasattr(app, "TRAIN_PROGRESS"):
            logger.debug("[SSE] 未检测到 TRAIN_PROGRESS，初始化为空字典")
            app.TRAIN_PROGRESS = {}

        user_xlsx = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'User.xlsx')
        try:
            logger.debug(f"[SSE] 读取用户信息文件: {user_xlsx}")
            df_user = pd.read_excel(user_xlsx)
            idx = df_user.index[df_user['userId'] == user_id].tolist()
            logger.debug(f"[SSE] 用户ID索引: {idx}")
            if not idx:
                logger.error(f"[SSE] 用户{user_id}不存在")
                yield f'data: {json.dumps({"error": "用户不存在"})}\n\n'
                return
            row_idx = idx[0]
            user_name = df_user.at[row_idx, 'name']
            ai_plan_str = df_user.at[row_idx, 'ai_plan']
            logger.debug(f"[SSE] 用户姓名: {user_name}, ai_plan: {ai_plan_str}")
            if not ai_plan_str or pd.isna(ai_plan_str):
                logger.warning(f"[SSE] 用户{user_id}没有训练计划")
                yield f'data: {json.dumps({"error": "该用户没有训练计划，请先进行初始评估"})}\n\n'
                return
            # 兼容字符串和字典
            if isinstance(ai_plan_str, dict):
                ai_plan = ai_plan_str
            else:
                try:
                    ai_plan = json.loads(ai_plan_str.replace("'", '"'))
                except Exception:
                    ai_plan = json.loads(ai_plan_str)
            actions = ai_plan.get('actions', [])
            logger.debug(f"[SSE] 训练动作列表: {actions}")
            for action in actions:
                if 'name' not in action:
                    action['name'] = get_action_name_by_id(action['action_id'])
        except Exception as e:
            logger.error(f"[SSE] 初始化训练进度失败: {str(e)}")
            yield f'data: {json.dumps({"error": f"初始化训练进度失败: {str(e)}"})}\n\n'
            return

        progress = app.TRAIN_PROGRESS.get(user_id)
        #if not progress:
        repeat = 2
        if actions and 'repeat' in actions[0]:
            repeat = int(actions[0]['repeat'])
        logger.info(f"[SSE] 初始化训练进度 repeat={repeat}")
        progress = {
            "actions": actions,
            "repeat": repeat,
            "current_round": 0,
            "current_action_index": 0,
            "results": [],
            "finished": False
        }
        app.TRAIN_PROGRESS[user_id] = progress

        actions = progress["actions"]
        repeat = progress["repeat"]
        round_idx = progress["current_round"]
        action_idx = progress["current_action_index"]
        results = progress["results"]

        logger.info(f"[SSE] 训练主循环开始，当前轮次: {round_idx}, 当前动作: {action_idx}")

        # === 3秒准备 ===
        prepare_total = 4
        prepare_start = time.time()
        for sec in range(prepare_total, 0, -1):
            now = time.time()
            seconds_left = int(prepare_total - (now - prepare_start))
            if seconds_left < 1:
                seconds_left = 1
            logger.debug(f"[SSE] 准备阶段，倒计时: {seconds_left}")
            yield f'data: {json.dumps({"status": "准备", "stage": "prepare", "seconds_left": seconds_left, "current_round": round_idx, "current_action_index": action_idx, "results": results})}\n\n'
            time.sleep(1)
        
        while True:
            logger.info(f"[SSE] 进入主循环，轮次: {round_idx}, 动作: {action_idx}")
            if progress.get("finished"):
                logger.info("[SSE] 训练过程完成，推送最终结果")
                yield f'data: {json.dumps({"status": "训练过程完成", "results": results})}\n\n'
                break
            if round_idx >= repeat:
                logger.info("[SSE] 所有轮次采集完成，标记finished")
                progress["finished"] = True
                yield f'data: {json.dumps({"status": "训练过程完成", "results": results})}\n\n'
                break

            action = actions[action_idx]
            action_id = action.get('action_id')
            action_name = action.get('name')
            logger.info(f"[SSE] 当前动作: {action_name} (action_id={action_id})")

            # === 采集前清空队列 ===
            logger.debug("[SSE] 采集前清空UDP数据队列")
            try:
                while not UDP_DATA_QUEUE.empty():
                    UDP_DATA_QUEUE.get_nowait()
            except Exception as e:
                logger.debug(f"[SSE] 清空队列异常: {e}")

            # === 采集10秒 ===
            collect_total = 10
            gesture_data = []
            collect_start = time.time()
            while True:
                now = time.time()
                elapsed = now - collect_start
                seconds_left = int(collect_total - elapsed + 0.999)
                if seconds_left < 0:
                    seconds_left = 0
                emg_snapshot = []
                start_loop = time.time()
                emg = None
                try:
                    while True:
                        emg_candidate = UDP_DATA_QUEUE.get_nowait()
                        if isinstance(emg_candidate, list) and len(emg_candidate) == 3:
                            gesture_data.append(emg_candidate)
                            emg_snapshot.append(emg_candidate)
                            logger.debug(f"[SSE] 采集到EMG数据: {emg_candidate}")
                except queue.Empty:
                    pass
                except Exception as e:
                    logger.debug(f"[SSE] 采集数据异常: {e}")
                logger.debug(f"[SSE] 采集中，剩余秒数: {seconds_left}, snapshot条数: {len(emg_snapshot)}")
                # === 修改：采集时也返回动作名和ID ===
                yield f'data: {json.dumps({"status": "采集中", "stage": "collect", "seconds_left": seconds_left, "current_round": round_idx, "current_action_index": action_idx, "action_name": action_name, "action_id": action_id, "realtime_emg": emg_snapshot, "results": results})}\n\n'
                sleep_time = 0.5 - (time.time() - start_loop)
                if sleep_time > 0:
                    time.sleep(sleep_time)
                if elapsed >= collect_total:
                    break

            logger.info(f"[SSE] 采集到数据条数: {len(gesture_data)}")
            arr = np.array(gesture_data)
            if arr.size > 0:
                rms, mnf, mf = calculate_emg_features(arr)
                logger.info(f"[SSE] 特征值: RMS={rms}, MNF={mnf}, MF={mf}")
            else:
                rms, mnf, mf = 0.0, 0.0, 0.0
                logger.warning("[SSE] 采集数据为空，特征值置为0")

            while len(results) <= round_idx:
                results.append([])
            results[round_idx].append({
                "action_id": action_id,
                "action_name": action_name,
                "data": gesture_data,
                "rms": rms,
                "mnf": mnf,
                "mf": mf
            })

            try:
                std_dir = os.path.join('user_files', str(user_name), 'training')
                os.makedirs(std_dir, exist_ok=True)
                std_file = os.path.join(std_dir, 'std.xlsx')
            
                rms_col = f"{action_id}_RMS"
                mnf_col = f"{action_id}_MNF"
                mf_col = f"{action_id}_MF"
                new_cols = [rms_col, mnf_col, mf_col]
            
                # 如果文件存在，读取；否则新建
                if os.path.exists(std_file):
                    df_std = pd.read_excel(std_file)
                    if df_std.empty:
                        df_std = pd.DataFrame(columns=new_cols)
                else:
                    df_std = pd.DataFrame(columns=new_cols)
            
                # 补全缺失列（如果没有）
                for col in new_cols:
                    if col not in df_std.columns:
                        df_std[col] = None
            
                # 确保至少有一行
                if df_std.shape[0] == 0:
                    df_std.loc[0] = [None] * len(df_std.columns)
            
                # 写入数据
                df_std.at[0, rms_col] = rms
                df_std.at[0, mnf_col] = mnf
                df_std.at[0, mf_col] = mf
            
                df_std.to_excel(std_file, index=False)
                logger.info(f"[SSE] 已写入/更新特征到 {std_file}: {rms_col}={rms}, {mnf_col}={mnf}, {mf_col}={mf}")
            
            except Exception as e:
                logger.error(f"[SSE] 写入std.xlsx失败: {e}")


            # === 休息5秒 ===
            rest_total = 6
            rest_start = time.time()
            for sec in range(rest_total, 0, -1):
                now = time.time()
                seconds_left = int(rest_total - (now - rest_start))
                if seconds_left < 1:
                    seconds_left = 1
                logger.debug(f"[SSE] 休息阶段，倒计时: {seconds_left}")
                yield f'data: {json.dumps({"status": "休息", "stage": "rest", "seconds_left": seconds_left, "current_round": round_idx, "current_action_index": action_idx, "results": results})}\n\n'
                time.sleep(1)

            # 推进动作或轮次
            if action_idx + 1 < len(actions):
                progress["current_action_index"] += 1
                logger.info(f"[SSE] 进入下一个动作: {progress['current_action_index']}")
            else:
                progress["current_action_index"] = 0
                progress["current_round"] += 1
                logger.info(f"[SSE] 进入下一个轮次: {progress['current_round']}")

            round_idx = progress["current_round"]
            action_idx = progress["current_action_index"]

    return Response(sse_stream(), mimetype='text/event-stream')

def send_feedback_to_ai(action_id, rms, mnf, mf, percent_avg , results, feedback_queue):
    global CURRENT_AI_SESSION  # ✅ 告诉Python：使用的是全局变量
    feedback = None

    if 'CURRENT_AI_SESSION' in globals() and CURRENT_AI_SESSION:
        payload = {
            "user": {
                "real_amount": {
                    "action_id": action_id,
                    "MF": mf,
                    "MNF": mnf,
                    "RMS": rms,
                    "PER_AVG": percent_avg
                }
            }
        }

        feedback = CURRENT_AI_SESSION.send_input(payload)
        
    feedback_queue.put(feedback)


def handle_feedback(feedback, results):
    """处理AI反馈并将其传递给前端"""
    logger.info(f"[SSE] AI反馈: {feedback}")  # 打印AI返回的结果
    # 将AI反馈传递给前端
    yield f'data: {json.dumps({"status": "AI反馈", "feedback": feedback, "results": results})}\n\n'

def process_data_and_train_model(data_dict, user_name):
    # 生成保存路径
    current_time = datetime.datetime.now().strftime("%Y-%m-%d")
    file_name = f"{current_time}_Training-data.xlsx"
    save_dir = os.path.join("user_files", user_name, "training", "Training_data")
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, file_name)

    # 保存原始数据（每个动作一个sheet，3列通道）
    with pd.ExcelWriter(file_path) as writer:
        for key, value in data_dict.items():
            arr = np.array(value)
            if arr.ndim == 3:
                arr = arr.reshape(-1, arr.shape[2])
            elif arr.ndim == 2:
                pass
            else:
                arr = np.array(value)
            df_emg = pd.DataFrame(arr, columns=[1, 2, 3])  # 三通道列头为1,2,3
            df_emg.to_excel(writer, sheet_name=key, index=False)

    logger.info(f"原始数据已保存到 {file_path}")

def save_features_to_excel(user_id: int, action_id: int, rms: float, mnf: float, mf: float) -> None:
    """
    将 (rms, mnf, mf) 写入 user_files/<name>/training/kangfu/data.xlsx 的【第 1 行】。
    若文件/目录/列不存在则自动创建。
    """
    user_name = get_user_name(user_id)
    if user_name is None:
        raise ValueError(f"无法找到 user_id={user_id} 的姓名，写入中止。")

    # 目录与文件路径
    base_dir   = os.path.join(CURRENT_DIR, "user_files", user_name, "training", "kangfu")
    os.makedirs(base_dir, exist_ok=True)
    excel_path = os.path.join(base_dir, "data.xlsx")

    # 列名
    col_rms, col_mnf, col_mf = f"{action_id}_RMS", f"{action_id}_MNF", f"{action_id}_MF"

    if os.path.exists(excel_path):
        # 读旧数据
        df = pd.read_excel(excel_path)

        # 若表为空先补一行
        if df.empty:
            df = pd.DataFrame([{}])

        # 补列
        for col in (col_rms, col_mnf, col_mf):
            if col not in df.columns:
                df[col] = pd.NA

        # 覆写第 0 行（Excel 第 1 行）
        df.loc[0, [col_rms, col_mnf, col_mf]] = [rms, mnf, mf]
    else:
        # 首次创建
        df = pd.DataFrame([{col_rms: rms, col_mnf: mnf, col_mf: mf}])

    # 保存
    df.to_excel(excel_path, index=False)


# SSE流：实时推送数据
@app.route('/api/predict', methods=['GET'])
def predict():
    user_id = request.args.get('id')
    logger.info(f"[SSE] 请求收到，用户ID: {user_id}")
    if not user_id:
        logger.error("[SSE] 未提供用户ID参数")
        return jsonify({"error": "需要提供id参数"}), 400

    def sse_stream():
        # 读取并初始化 User.xlsx
        user_xlsx_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'User.xlsx')
        try:
            logger.info(f"[SSE] 读取用户信息文件: {user_xlsx_path}")
            df_user = pd.read_excel(user_xlsx_path)
            percent_cols = [c for c in df_user.columns if c.startswith('percent')]
            for c in percent_cols:
                df_user[c] = df_user[c].astype('float64')

            idx = df_user.index[df_user['userId'] == user_id].tolist()
            logger.info(f"[SSE] 用户ID索引: {idx}")
            if not idx:
                yield f'data: {json.dumps({"error": "用户不存在"})}\n\n'
                return
            row_idx = idx[0]
            user_name = df_user.at[row_idx, 'name']
            ai_plan_str = df_user.at[row_idx, 'ai_plan']
            logger.info(f"[SSE] 用户姓名: {user_name}, ai_plan: {ai_plan_str}")
            if not ai_plan_str or pd.isna(ai_plan_str):
                yield f'data: {json.dumps({"error": "该用户没有训练计划，请先进行初始评估"})}\n\n'
                return

            # 解析训练计划
            if isinstance(ai_plan_str, dict):
                ai_plan = ai_plan_str
            else:
                ai_plan = json.loads(ai_plan_str.replace("'", '"'))
            actions = ai_plan.get('actions', [])
            logger.info(f"[SSE] 训练动作列表: {actions}")
            for action in actions:
                if 'name' not in action:
                    action['name'] = get_action_name_by_id(action['action_id'])
                logger.info(f"[SSE] 动作名称: {action['name']}, 动作ID: {action['action_id']}")
        except Exception as e:
            logger.error(f"[SSE] 初始化预测进度失败: {e}")
            yield f'data: {json.dumps({"error": f"初始化预测进度失败: {e}"})}\n\n'
            return

        # <<< prep stage: 五秒准备倒计时 >>>
        if actions:
            next_action_name = actions[0]['name']
        else:
            next_action_name = None

        for sec in range(5, 0, -1):
            payload_prep = {
                "status": "准备",
                "stage": "rest",
                "seconds_left": sec,
                "next_action_name": next_action_name
            }
            yield 'data: ' + json.dumps(payload_prep) + '\n\n'
            time.sleep(1)
        # <<< end prep stage >>>

        results = []
        raw_data_dict = {}

        # 主训练循环
        for action_idx, action in enumerate(actions):
            action_id = action.get('action_id')
            action_name = action.get('name', f"动作{action_id}")
            repeat_count = action.get('repeat', 1)
            time_seconds = int(action.get('time', 10))
            logger.info(f"[SSE] 动作 {action_name} 开始，重复次数: {repeat_count}，训练时长: {time_seconds}秒")

            all_percent_rms, all_percent_mnf, all_percent_mf = [], [], []
            action_gesture_data = []

            for round_idx in range(repeat_count):
                logger.info(f"[SSE] 第{round_idx+1}轮训练：{action_name}")
                gesture_data = []
                collect_start = time.time()
                while True:
                    elapsed = time.time() - collect_start
                    seconds_left = max(int(time_seconds - elapsed + 0.999), 0)

                    emg_snapshot = []
                    loop_start = time.time()
                    try:
                        while True:
                            emg_candidate = UDP_DATA_QUEUE.get_nowait()
                            if isinstance(emg_candidate, list) and len(emg_candidate) == 3:
                                gesture_data.append(emg_candidate)
                                emg_snapshot.append(emg_candidate)
                    except queue.Empty:
                        pass

                    yield f'data: {json.dumps({ "status": "采集中", "action_name": action_name, "realtime_emg": emg_snapshot, "seconds_left": seconds_left, "round_idx": round_idx+1, "repeat": repeat_count })}\n\n'

                    time.sleep(max(0, 0.5 - (time.time() - loop_start)))
                    if elapsed >= time_seconds:
                        break

                arr = np.array(gesture_data)
                if arr.size > 0:
                    rms, mnf, mf = calculate_emg_features(arr)
                else:
                    rms = mnf = mf = 0.0

                save_features_to_excel(user_id, action_id, rms, mnf, mf)

                std_file = os.path.join('user_files', str(user_name), 'training', 'std.xlsx')
                if os.path.exists(std_file):
                    std_df = pd.read_excel(std_file)
                else:
                    std_df = pd.DataFrame(columns=[f"{action_id}_{m}" for m in ('RMS','MNF','MF')])
                    logger.warning("[SSE] 标准值文件不存在，已创建默认标准值")

                std_rms = std_df.at[0, f"{action_id}_RMS"] if f"{action_id}_RMS" in std_df.columns else None
                std_mnf = std_df.at[0, f"{action_id}_MNF"] if f"{action_id}_MNF" in std_df.columns else None
                std_mf  = std_df.at[0, f"{action_id}_MF"]  if f"{action_id}_MF"  in std_df.columns else None

                percent_rms = (rms / std_rms * 100) if std_rms else None
                percent_mnf = (mnf / std_mnf * 100) if std_mnf else None
                percent_mf  = (mf  / std_mf  * 100) if std_mf  else None

                all_percent_rms.append(percent_rms)
                all_percent_mnf.append(percent_mnf)
                all_percent_mf.append(percent_mf)
                action_gesture_data.extend(gesture_data)

                # 轮间休息
                if round_idx < repeat_count - 1:
                    for rest in range(10, 0, -1):
                        yield f'data: {json.dumps({ "status": "休息", "stage": "short_rest" , "seconds_left": rest, "action_name": action_name })}\n\n'
                        time.sleep(1)

            # 汇总并写回 User.xlsx
            avg_r = sum(all_percent_rms) / len(all_percent_rms)
            avg_n = sum(all_percent_mnf) / len(all_percent_mnf)
            avg_f = sum(all_percent_mf)  / len(all_percent_mf)
            results.append({
                "action_id": action_id,
                "action_name": action_name,
                "rms": rms, "mnf": mnf, "mf": mf,
                "avg_percent_rms": avg_r,
                "avg_percent_mnf": avg_n,
                "avg_percent_mf":  avg_f
            })

            try:
                percent_avg = (avg_r + avg_n + avg_f) / 3
                col_name = f"percent{action_idx+1}"
                df_write = pd.read_excel(user_xlsx_path)
                if col_name not in df_write.columns:
                    df_write[col_name] = 0.0
                df_write[col_name] = df_write[col_name].astype('float64')
                uidxs = df_write.index[df_write['userId'] == user_id].tolist()
                if uidxs:
                    df_write.at[uidxs[0], col_name] = float(percent_avg)
                    df_write.to_excel(user_xlsx_path, index=False)
                    logger.info(f"[SSE] 写入 {col_name} = {percent_avg:.2f} 到 User.xlsx")
            except Exception as e:
                logger.error(f"[SSE] 写入 {col_name} 失败: {e}")

            # AI 反馈线程
            feedback_queue = queue.Queue()
            ai_thread = threading.Thread(
                target=send_feedback_to_ai,
                args=(action_id, avg_r, avg_n, avg_f, percent_avg, results, feedback_queue)
            )
            ai_thread.start()

            # 动作间休息 + AI 反馈
            if action_idx < len(actions) - 1:
                next_action_name = actions[action_idx + 1]['name']
                rest_total = 20
                rest_start = time.time()
                while True:
                    left = int(rest_total - (time.time() - rest_start))
                    # AI 反馈
                    if not feedback_queue.empty():
                        fb = feedback_queue.get()
                        yield 'data: ' + json.dumps({
                            "status": "AI反馈",
                            "feedback": fb,
                            "results": results
                        }) + '\n\n'
                    if left > 0:
                        yield 'data: ' + json.dumps({
                            "status": "休息",
                            "stage": "rest",
                            "seconds_left": left,
                            "next_action_name": next_action_name,
                            "results": results
                        }) + '\n\n'
                        time.sleep(1)
                        continue
                    break

            if action_idx == len(actions) - 1:
                rest_total = 20
                rest_start = time.time()
                while True:
                    left = int(rest_total - (time.time() - rest_start))
                    # AI 反馈
                    if not feedback_queue.empty():
                        fb = feedback_queue.get()
                        yield 'data: ' + json.dumps({
                            "status": "AI反馈",
                            "feedback": fb,
                            "results": results
                        }) + '\n\n'
                    if left > 0:
                        yield 'data: ' + json.dumps({
                            "status": "休息",
                            "stage": "rest",
                            "seconds_left": left,
                            "next_action_name": "所有训练已结束，请准备康复后的肌肉评估",
                            "results": results
                        }) + '\n\n'
                        time.sleep(1)
                        continue
                    break

            while not feedback_queue.empty():
                feedback_queue.get_nowait()

            raw_data_dict[action_name] = action_gesture_data
        
        process_data_and_train_model(raw_data_dict, user_name)
        logger.info("[SSE] 预测过程完成")
        yield f'data: {json.dumps({"status": "预测过程完成", "results": results})}\n\n'

    headers = {
        "Content-Type":             "text/event-stream",
        "Cache-Control":            "no-cache",
        "Connection":               "keep-alive",
        "Access-Control-Allow-Origin": "*"
    }
    return Response(sse_stream(), headers=headers)



@app.route('/api/udp-emg-stream')
def udp_emg_stream():
    """实时推送UDP采集到的EMG数据（SSE流）"""
    def event_stream():
        import time
        while True:
            try:
                emg = UDP_DATA_QUEUE.get(timeout=5)
                #print(f"[SSE] 发送EMG数据到前端: {emg}")  # 调试用
                yield f"data: {json.dumps({'emg': emg})}\n\n"
            except Exception:
                yield "data: {}\n\n"
    return Response(event_stream(), mimetype='text/event-stream')
    
@app.route('/api/search_doctors', methods=['GET'])
def search_doctors():
    """
    根据关键字模糊搜索 Doctor.xlsx 中的医生姓名，
    返回符合条件的列表，每项包含 name, major, school, hospital, description。
    """
    keyword = request.args.get('keyword', '').strip()
    if not keyword:
        return jsonify({"doctors": []})

    base_dir = os.path.abspath(os.path.dirname(__file__))
    doctor_xlsx = os.path.join(base_dir, 'Doctor.xlsx')
    if not os.path.exists(doctor_xlsx):
        return jsonify({"error": "Doctor.xlsx 文件不存在"}), 500

    # 读取医生信息
    df_doc = pd.read_excel(doctor_xlsx, engine='openpyxl')

    # 模糊搜索医生姓名
    mask = df_doc['name'].str.contains(keyword, case=False, na=False)
    matches = df_doc.loc[mask, ['id', 'name', 'major', 'hospital', 'description']]  # 添加 hospital 和 description 列

    # 转换为字典格式并返回
    doctors = matches.to_dict(orient='records')
    return jsonify({"doctors": doctors})


@app.route('/api/feature-lookup', methods=['POST'])
def feature_lookup():
    """
    接收 id, name, age, gender, ache, ai_name, doctor_name。
    更新 User.xlsx、Features.xlsx，然后把 user_id 写入对应 doctor 的 patient_list.json。
    """
    data = request.get_json()
    user_id     = data.get('id')
    name        = data.get('name')
    age         = data.get('age')
    gender_sent = data.get('gender')
    ache        = data.get('ache')
    ai_name     = data.get('ai_name')
    doctor_name = data.get('doctor_name')
    device_id   = data.get('device_id')

    # 参数校验
    if None in (user_id, name, age, gender_sent, ache, ai_name, doctor_name):
        return jsonify({"error": "缺少参数"}), 400

    # 性别映射
    if gender_sent.lower() == 'male':
        gender_text = '男'
    elif gender_sent.lower() == 'female':
        gender_text = '女'
    else:
        gender_text = gender_sent

    base_dir        = os.path.abspath(os.path.dirname(__file__))
    user_data_path  = os.path.join(base_dir, 'User.xlsx')
    features_path   = os.path.join(base_dir, 'Features.xlsx')
    doctor_xlsx     = os.path.join(base_dir, 'Doctor.xlsx')

    try:
        # —— User.xlsx 初始化或读取 —— #
        if not os.path.exists(user_data_path):
            df_init = pd.DataFrame(columns=[ 
                "userId", "name", "age", "gender", "ache",
                "1_RMS_std", "1_MNF_std", "1_MF_std",
                "2_RMS_std", "2_MNF_std", "2_MF_std",
                "AI_name", "api_key", "doctor", "device_id"  # ✅ 加上这个”  # 添加 doctor 列
            ])
            df_init.to_excel(user_data_path, index=False)

        df = pd.read_excel(user_data_path, engine='openpyxl')

        # 重名冲突检查
        df['userId'] = df['userId'].astype(str)

        conflict = df[ 
            (df['name'] == name) & 
            (df['userId'] != user_id)
        ]
        if not conflict.empty:
            return jsonify({"error": "此姓名已注册"}), 409
            
        if 'device_id' not in df.columns:
            df['device_id'] = None

        # 更新或新增用户基本信息
        if user_id in df['userId'].values:
            df.loc[df['userId'] == user_id, ['name','age','gender','ache','device_id']] = [ 
                name, age, gender_text, ache, device_id
            ]
        else:
            new_row = pd.DataFrame({
                "userId": [user_id],
                "name": [name],
                "age": [age],
                "gender": [gender_text],
                "ache": [ache],
                "doctor": [None],  # 确保创建 doctor 列
                "device_id":[device_id]
            })
            df = pd.concat([df, new_row], ignore_index=True)

        # —— Features.xlsx 处理 —— #
        if not os.path.exists(features_path):
            return jsonify({"error": "Features.xlsx 文件不存在"}), 500
        df_feat = pd.read_excel(features_path, engine='openpyxl')
        df_feat['Gender'] = df_feat['Gender'].astype(str)
        df_feat['age'] = df_feat['age'].astype(int)

        matched = df_feat[
            (df_feat['Gender'] == gender_sent) &
            (df_feat['age'] == int(age))
        ]
        if matched.empty:
            return jsonify({"error": "未找到匹配的特征数据"}), 404
        row = matched.iloc[0]

        df.loc[df['userId'] == user_id, [
            '1_RMS_std','1_MNF_std','1_MF_std'
        ]] = [row['1_RMS_mean'], row['1_MNF_mean'], row['1_MF_mean']]
        df.loc[df['userId'] == user_id, [
            '2_RMS_std','2_MNF_std','2_MF_std'
        ]] = [row['2_RMS_mean'], row['2_MNF_mean'], row['2_MF_mean']]

        # —— 选择 AI 并生成 API Key —— #
        AI_MODELS = {
            "DeepSeek": {
                "name": "DeepSeek",
                "session_class": DeepSeekSession,
                "default_key": "sk-24fa19bacfaa41c1a9c70bbf98b3800f"
            },
            "通义千问": {
                "name": "通义千问",
                "session_class": TongyiSession,
                "default_key": "sk-3b2b5fd8fd2e42fc8d53541b595095b5"
            },
            "豆包": {
                "name": "豆包",
                "session_class": VolcanoSession,
                "default_key": "c6a25fc5-a553-4eb4-9d3d-82e362b0c6ce"
            }
        }

        ai_model = AI_MODELS.get(ai_name)
        if ai_model is None:
            return jsonify({"error": "未知的 AI 名称"}), 404

        api_key = ai_model["default_key"]
        global CURRENT_AI_SESSION
        CURRENT_AI_SESSION = ai_model["session_class"](api_key)

        if 'api_key' not in df.columns:
            df['api_key'] = None
        df.loc[df['userId'] == user_id, 'api_key'] = api_key

        # 保存 User.xlsx
        df.to_excel(user_data_path, index=False)

        # —— 更新 doctor 列 —— #
        # 获取医生的名字
        doctor_column = df.columns.get_loc("doctor")
        current_doctor_value = df.loc[df['userId'] == user_id, 'doctor'].values[0]
        
        if not current_doctor_value:
            # 如果该单元格为空，直接填入医生名字
            df.loc[df['userId'] == user_id, 'doctor'] = doctor_name
        else:
            # 如果已有医生名字，则检查是否已经存在该医生
            doctor_names = current_doctor_value.split(", ")
            if doctor_name not in doctor_names:
                doctor_names.append(doctor_name)
                df.loc[df['userId'] == user_id, 'doctor'] = ", ".join(doctor_names)
            else:
                return jsonify({"message": "该医生已存在", "doctor_name": doctor_name}), 200

        # 保存更新后的 User.xlsx
        df.to_excel(user_data_path, index=False)

        # —— 将 user_id 写入对应医生的 patient_list.json —— #
        if not os.path.exists(doctor_xlsx):
            return jsonify({"error": "Doctor.xlsx 文件不存在"}), 500
        df_doc = pd.read_excel(doctor_xlsx, engine='openpyxl')
        match = df_doc[df_doc['name'] == doctor_name]
        if match.empty:
            return jsonify({"error": f"未找到医生 {doctor_name}"}), 404
        doctor_id = str(match.iloc[0]['id'])

        doctor_folder = os.path.join(base_dir, 'user_files', doctor_id)
        os.makedirs(doctor_folder, exist_ok=True)
        plist_path = os.path.join(doctor_folder, 'patient_list.json')

        if os.path.exists(plist_path):
            with open(plist_path, 'r', encoding='utf-8') as f:
                patient_list = json.load(f)
        else:
            patient_list = []

        if user_id not in patient_list:
            patient_list.append(user_id)
            with open(plist_path, 'w', encoding='utf-8') as f:
                json.dump(patient_list, f, ensure_ascii=False, indent=2)

        # 返回用户信息以及医生的名字
        user_doctor = df.loc[df['userId'] == user_id, 'doctor'].values[0]

        return jsonify({
            "success": True,
            "message": "修改成功",
            "doctor_name": user_doctor  # 返回医生的名字
        })

    except Exception as e:
        return jsonify({"error": f"内部错误: {e}"}), 500


#获取当前医生列表
@app.route('/api/get_doctor', methods=['POST'])
def get_doctor():
    """
    根据用户的 id 返回对应的医生名字
    接收参数:
    {
        "id": "用户的id"
    }
    返回：
    {
        "doctor_name": "医生的名字" 或 "当前没有医生信息"
    }
    """
    data = request.get_json()
    user_id = data.get("id")

    if not user_id:
        return jsonify({"error": "缺少id参数"}), 400

    user_data_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "User.xlsx")

    if not os.path.exists(user_data_path):
        return jsonify({"error": "User.xlsx 文件不存在"}), 500

    # 读取 User.xlsx 文件
    try:
        df = pd.read_excel(user_data_path, engine="openpyxl")
    except Exception as e:
        return jsonify({"error": f"读取 User.xlsx 文件失败: {str(e)}"}), 500

    # 查找用户的 doctor 信息
    user_row = df[df["userId"] == str(user_id)]

    if user_row.empty:
        return jsonify({"error": "未找到该用户"}), 404

    # 获取 doctor 列的值
    doctor_name = user_row["doctor"].values[0]

    if not doctor_name:
        return jsonify({"doctor_name": "当前没有医生信息"}), 200
    else:
        return jsonify({"doctor_name": doctor_name}), 200

#//////////////////////////////////////////////////////////////////////////////////////////
#获取医生信息
def get_all_doctors_info():
    try:
        workbook = openpyxl.load_workbook('Doctor.xlsx')
        sheet = workbook.active  # 获取活动的sheet

        doctors_info = []
        
        # 从第二行开始遍历（假设第一行是标题）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 假设第1列是id，第2列是name，依此类推
            doc_id, name, major, school, hospital, description, detailed_description = row
            doctors_info.append({
                "name": name,
                "major": major,
                "school": school,
                "hospital": hospital,
                "description": description,
                "detailedDescription": detailed_description
            })
        
        return doctors_info
    except Exception as e:
        print(f"Error reading Doctor.xlsx: {e}")
        return []

# 定义后端接口
@app.route('/api/find_doctor', methods=['POST'])
def find_doctor():
    # 获取前端传入的用户id
    data = request.get_json()
    user_id = data.get('id')

    if not user_id:
        return jsonify({"error": "缺少用户ID"}), 400

    # 获取所有医生信息
    doctors_info = get_all_doctors_info()
    
    if doctors_info:
        return jsonify({"doctors": doctors_info})
    else:
        return jsonify({"error": "未找到医生信息"}), 404
        
@app.route('/api/init-ai-session', methods=['POST'])
def init_ai_session():
    """
    前端发送 user_id。
    后端根据 User.xlsx 中的 api_key 初始化对应的 AI 模型并创建 CURRENT_AI_SESSION。
    若该用户没有 api_key，则默认使用 DeepSeek，并写入文件。
    返回最终使用的 api_key。
    """
    data = request.get_json()
    user_id = data.get('id')
    global CURRENT_AI_SESSION  # Make sure this is accessible globally

    if not user_id:
        return jsonify({"error": "缺少用户 ID"}), 400

    user_data_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'User.xlsx')

    AI_MODELS = {
        "DeepSeek": {
            "name": "DeepSeek",
            "session_class": DeepSeekSession,
            "default_key": "sk-24fa19bacfaa41c1a9c70bbf98b3800f"
        },
        "通义千问": {
            "name": "通义千问",
            "session_class": TongyiSession,
            "default_key": "sk-3b2b5fd8fd2e42fc8d53541b595095b5"
        },
        "豆包": {
            "name": "豆包",
            "session_class": VolcanoSession,
            "default_key": "c6a25fc5-a553-4eb4-9d3d-82e362b0c6ce"
        }
    }

    try:
        if not os.path.exists(user_data_path):
            return jsonify({"error": "用户数据文件不存在"}), 500

        df = pd.read_excel(user_data_path, engine='openpyxl')

        if user_id not in df['userId'].values:
            return jsonify({"error": "用户不存在"}), 404

        # 获取该用户的 api_key
        api_key = df.loc[df['userId'] == user_id, 'api_key'].values[0] if 'api_key' in df.columns else None

        # 匹配模型
        model_found = False
        session_class = None

        if api_key and isinstance(api_key, str) and api_key.strip() != "":
            for model in AI_MODELS.values():
                if api_key == model["default_key"]:
                    session_class = model["session_class"]
                    logger.info(f"[SSE] 选择的AI模型: {model['name']}, 使用API密钥: {api_key.strip()}")
                    model_found = True
                    break

        # 如果未找到模型或 api_key 为空，则默认使用 DeepSeek
        if not model_found:
            default_model = AI_MODELS["DeepSeek"]
            api_key = default_model["default_key"]
            session_class = default_model["session_class"]
            df.loc[df['userId'] == user_id, 'api_key'] = api_key  # 写入默认值
            df.to_excel(user_data_path, index=False)

        # 初始化会话
        CURRENT_AI_SESSION = session_class(api_key)

        return jsonify({"api_key": api_key})

    except Exception as e:
        logger.error(f"[ERROR] 初始化 AI 模型失败: {str(e)}")
        return jsonify({"error": f"初始化 AI 模型失败: {str(e)}"}), 500
        
#///////////////////////////////////////////////////////////////////////////////////////////////////////////
#获取coze资源
# ==== 读取私钥 ====
with open(private_key_path, "r") as f:
    private_key = f.read()

# ==== 存储有效的客户端配置 ====
valid_clients = [
    {
        "client_id": "1136790881240",
        "client_name": "Client A",
        "allowed_api": ["/api/init-ai-session", "/api/get_access_token"]
    }
]

# 检查客户端ID及权限
def check_client_id(client_id, requested_api):
    for client in valid_clients:
        if client["client_id"] == client_id:
            if requested_api in client["allowed_api"]:
                return True
            else:
                return False  # 客户端没有权限访问此 API
    return False  # client_id 无效

# ==== 构造 JWT ====
def generate_jwt(session_name):
    import jwt
    logger.info("Starting JWT generation...")
    
    # 获取当前时间戳
    issued_at = int(time.time())
    logger.info(f"Current time (issued_at): {issued_at}")
    
    # 构造 JWT payload
    jwt_payload = {
        "iss": client_id,
        "aud": api_base.replace("https://", ""),  # 只要主机域名
        "iat": issued_at,
        "exp": issued_at + 3600,  # Token有效期为1小时
        "jti": str(uuid.uuid4()),
        "session_name": session_name
    }
    
    # 构造 JWT headers，确保 kid 是正确的
    jwt_headers = {
        "kid": public_key_id  # 确保在这里传递正确的 kid
    }
    logger.info(f"Verifying JWT with client_id: {client_id}")
    logger.info(f"JWT Payload: {jwt_payload}")
    logger.info(f"JWT Headers: {jwt_headers}")

    try:
        # 使用 flask_jwt_extended 的 create_access_token 生成 JWT
        token = jwt.encode(jwt_payload, private_key, algorithm="RS256", headers=jwt_headers)
        logger.info("JWT successfully generated")
        logger.info(f"Generated JWT: {token[:20]}...")  # 输出 token 的一部分，以便验证，避免输出完整 token 过长
    except Exception as e:
        logger.error(f"Error generating JWT: {str(e)}")
        raise

    return token

# ==== 请求 Access Token ====
def request_access_token(jwt_token):
    headers = {
        "Authorization": f"Bearer {jwt_token}",
        "Content-Type": "application/json"
    }

    data = {
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
        "duration_seconds": 3600  # Token有效期为1小时
    }

    res = requests.post(token_url, headers=headers, json=data)

    # 检查请求是否成功
    if res.status_code != 200:
        logger.info(f"Error: {res.status_code}")
        try:
            error_response = res.json()
            logger.info(f"Error Message: {error_response.get('error_message', 'No error message')}")
            logger.info(f"Error Code: {error_response.get('error_code', 'No error code')}")
        except Exception as e:
            logger.info(f"Failed to parse error response: {str(e)}")
        return None  # Return None if the request fails
    else:
        access_token = res.json().get("access_token")
        logger.info(f"Access Token: {access_token}")
        expires_in = res.json().get("expires_in", "No expiration info")
        logger.info(f"Expires In: {expires_in} seconds")
        return access_token  # Return the access token


@app.route('/api/get_access_token', methods=['POST'])
def get_access_token():
    # 获取用户传入的ID
    user_id = request.json.get('id', None)

    if not user_id:
        return jsonify({"error": "User ID is required"}), 400

    # 验证 client_id 和权限
    if not check_client_id(client_id, "/api/get_access_token"):
        return jsonify({"error": "invalid client"}), 401

    # 生成JWT
    jwt_token = generate_jwt(session_name=user_id)
    
    logger.info(f"jwt_token:{jwt_token}")

    # 请求Access Token
    access_token = request_access_token(jwt_token)  # Now capture the access_token returned from request_access_token

    # 返回生成的access_token给前端
    if access_token:
        return jsonify({"access_token": access_token}), 200
    else:
        return jsonify({"error": "Failed to retrieve access token"}), 500
        

#///////////////////////////////////////////////////////////////////////
#RTC tokenId与AppId的获取
def generate_rtc_token(room_id, user_id, expire_seconds=3600):
    """生成RTC Token，正确引用权限常量"""
    try:
        # 实例化AccessToken类
        token = RTCAccessToken(
            RTC_APP_ID, 
            RTC_APP_KEY, 
            room_id, 
            user_id
        )
        
        current_ts = int(time.time())
        expire_ts = current_ts + expire_seconds
        
        # 关键修复：从模块级别引用权限常量
        # 这些常量定义在AccessToken.py的全局作用域中，而非类内部
        token.add_privilege(AccessToken.PrivPublishStream, expire_ts)
        token.add_privilege(AccessToken.PrivSubscribeStream, expire_ts)
        
        # 设置过期时间
        token.expire_time(expire_ts)
        
        generated_token = token.serialize()
        # 保证为字符串类型
        if isinstance(generated_token, bytes):
            generated_token = generated_token.decode('utf-8')
        logger.info(f"Token生成成功 - 长度: {len(generated_token)}, 前缀: {generated_token[:3]}")
        logger.info(f"Token repr: {repr(generated_token)}")
        return generated_token
        
    except Exception as e:
        logger.error(f"生成RTC Token失败: {str(e)}", exc_info=True)
        return None


@app.route('/api/get_rtc_auth', methods=['POST'])
def get_rtc_auth():
    try:
        data = request.get_json()
        
        if not data or 'userId' not in data or 'roomId' not in data:
            return jsonify({
                'success': False,
                'message': '缺少必要参数(userId或roomId)'
            }), 400
        
        user_id = data['userId']
        room_id = data['roomId']
        
        # 验证参数
        if not isinstance(user_id, str) or not user_id.strip():
            return jsonify({'success': False, 'message': 'userId必须是非空字符串'}), 400
        if not isinstance(room_id, str) or not room_id.strip():
            return jsonify({'success': False, 'message': 'roomId必须是非空字符串'}), 400
        if len(RTC_APP_ID) != 24:
            return jsonify({'success': False, 'message': f'AppID长度错误（应为24位，实际{len(RTC_APP_ID)}位）'}), 500
        
        rtc_token = generate_rtc_token(room_id, user_id, 3600)
        if not rtc_token:
            return jsonify({'success': False, 'message': '生成RTC Token失败'}), 500

        # 保证token为str类型
        if isinstance(rtc_token, bytes):
            rtc_token = rtc_token.decode('utf-8')

        return jsonify({
            'success': True,
            'appId': RTC_APP_ID,
            'token': rtc_token,
            'roomId': room_id,
            'userId': user_id,
            'expiresIn': 3600
        })
        
    except Exception as e:
        logger.error(f"处理RTC认证请求出错: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500
        
@app.route('/api/manual_plan', methods=['POST'])
def manual_plan():
    try:
        data = request.get_json()  # 获取请求体中的 JSON 数据
        user_id = data.get("user_id")  # 获取 user_id
        actions_raw = data.get("actions")  # 获取动作列表 [{action_id: ..., time: ...}]

        # 参数校验
        if not user_id or not actions_raw or not isinstance(actions_raw, list):
            return jsonify({"error": "缺少 user_id 或 actions"}), 400

        # 动作名映射
        action_names = {
            1: '握拳与打开手掌',
            2: '手掌旋转',
            3: '腕屈曲',
            4: '腕伸展',
            5: '手心向自己，手掌向内侧旋转',
            6: '手心向自己，手掌向外侧旋转',
            7: '压手'
        }

        # 功能解释模板
        training_purposes = {
            1: "增强手指和腕部协同控制能力",
            2: "提升手掌的旋转灵活性",
            3: "强化腕屈曲肌群力量",
            4: "改善腕关节伸展活动度",
            5: "提升内旋时的稳定控制",
            6: "提升外旋时的灵活性",
            7: "增强腕部下压力量与耐力"
        }

        # 构建动作列表描述
        action_phrases = []
        training_goals = []

        for item in actions_raw:
            aid = item["action_id"]  # 动作 ID
            t = item["time"]  # 时间
            name = action_names.get(aid, f"动作{aid}")  # 获取动作名称
            goal = training_purposes.get(aid, f"{name} 的功能训练")  # 获取功能训练目标

            # 构建动作描述
            action_phrases.append(f"{name} {t}秒")
            training_goals.append(goal)

        # 将所有动作和目标组合成字符串
        action_text = "，".join(action_phrases)
        goal_text = "，".join(training_goals)

        # 创建最终的消息内容
        message = (
            f"今天的康复计划是：{action_text}。结合您的训练状态，这些动作可以帮助您逐步恢复功能。"
            f"其目的包括：{goal_text}，从而实现更全面的肌肉控制和关节稳定性提升。"
        )

        # 计划内容
        plan = {
            "type": "plan",
            "actions": actions_raw,
            "message": message
        }

        # 写入 Excel 或数据库（你需要实现这个函数）
        update_ai_plan_in_excel(user_id, plan)

        return jsonify({"status": "ok", "plan": plan})

    except Exception as e:
        # 捕获异常并记录错误
        logger.error(f"[ManualPlan] 创建手动康复计划失败: {e}")
        return jsonify({"error": str(e)}), 500



#///////////////////////////////////////////////////////////////////////////////////////////////
# app/report_api.py

def read_user_row(user_id: int) -> pd.Series:
    """
    按 id 读取 User.xlsx 中的整行记录
    """
    logger.info("读取用户信息: id=%s", user_id)
    df = pd.read_excel(USER_FILE, engine="openpyxl")

    if "userId" not in df.columns:
        raise KeyError("User.xlsx 必须包含 'id' 列")

    row = df.loc[df["userId"] == user_id]
    if row.empty:
        raise ValueError(f"未找到 id={user_id} 的用户")

    return row.iloc[0]


def latest_file(pattern: str) -> str:
    """
    返回匹配 pattern 的最新文件路径
    """
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"未找到文件: {pattern}")

    newest = max(files, key=os.path.getmtime)
    logger.info("最新文件: %s", newest)
    return newest


def build_rehab_training(user_dir: str, user_row: pd.Series) -> list[dict[str, any]]:
    """
    生成 rehab_training_data 列表
    只会取与 User.xlsx 中 id 对应的 ai_plan 中包含的动作 id
    """
    path = os.path.join(user_dir, "training", "kangfu", "data.xlsx")
    logger.info("读取康复训练数据: %s", path)
    df = pd.read_excel(path, engine="openpyxl")

    # 假设 user_row 中有 "id" 字段，代表当前用户的 id
    user_id = user_row['userId']
    
    # 读取 User.xlsx 文件来获取与 id 相关的 ai_plan 数据
    user_xlsx_path = os.path.join('User.xlsx')
    df_user = pd.read_excel(user_xlsx_path, engine="openpyxl")

    # 获取当前用户的 ai_plan 数据（假设 "ai_plan" 是一列，其中存储了该用户的训练动作计划）
    user_ai_plan = df_user[df_user['userId'] == user_id]['ai_plan'].values

    # 如果 ai_plan 为空或找不到对应用户的数据，则返回空列表
    if len(user_ai_plan) == 0:
        logger.warning(f"未找到用户 {user_id} 的 ai_plan 数据")
        return []

    # ai_plan 数据是字典形式，我们需要从中提取 "actions" 列表
    ai_plan_data = eval(user_ai_plan[0])  # 将字符串转换为字典
    actions = ai_plan_data['actions']

    records: list[dict[str, any]] = []
    
    # 遍历用户的 ai_plan 中的每个动作，根据 action_id 获取数据
    for action in actions:
        action_id = action['action_id']
        
        # 遍历 data.xlsx 的每一列，提取与 action_id 对应的数据
        rms_col = f"{action_id}_RMS"
        if rms_col not in df.columns:
            continue  # 如果没有对应的 RMS 列，则跳过该 action_id
        
        rms = float(df[rms_col].iloc[-1])
        mnf = float(df[f"{action_id}_MNF"].iloc[-1])
        mf = float(df[f"{action_id}_MF"].iloc[-1])

        # 获取用户的百分比数据
        percent_key = f"percent{len(records) + 1}"
        percent = float(user_row.get(percent_key, 1.0))

        logger.debug("动作 %s → MF=%s, MNF=%s, RMS=%s, percent=%s",
                     action_id, mf, mnf, rms, percent)

        # 将当前动作的数据添加到记录中
        records.append({
            "action_id": action_id,
            "MF": mf,
            "MNF": mnf,
            "RMS": rms,
            "percent": percent
        })

    logger.info("rehab_training_data 共 %d 条", len(records))
    return records



def get_ai_session(user_id: int) -> tuple[object, str]:
    """
    根据 User.xlsx 的 api_key 匹配 AI 会话；无则 fallback 到 DeepSeek。
    会话对象存入 CURRENT_AI_SESSION 复用。
    """
    global CURRENT_AI_SESSION
    
        # 防御：若被意外覆盖为非 dict，则重置
    if not isinstance(CURRENT_AI_SESSION, dict):
        logging.warning("CURRENT_AI_SESSION 被覆盖为 %s，已重置为空字典",
                        type(CURRENT_AI_SESSION))
        CURRENT_AI_SESSION = {}

    # 可选模型
    AI_MODELS = {
        "DeepSeek": {
            "name": "DeepSeek",
            "session_class": DeepSeekSession,
            "default_key": "sk-24fa19bacfaa41c1a9c70bbf98b3800f"
        },
        "通义千问": {
            "name": "通义千问",
            "session_class": TongyiSession,
            "default_key": "sk-3b2b5fd8fd2e42fc8d53541b595095b5"
        },
        "豆包": {
            "name": "豆包",
            "session_class": VolcanoSession,
            "default_key": "c6a25fc5-a553-4eb4-9d3d-82e362b0c6ce"
        }
    }

    # 读取用户
    df = pd.read_excel(USER_FILE, engine="openpyxl")
    if user_id not in df["userId"].values:
        raise ValueError("用户不存在")

    api_key = str(df.loc[df["userId"] == user_id, "api_key"].values[0]) if "api_key" in df.columns else ""

    # 选择模型
    session_class = None
    for model in AI_MODELS.values():
        if api_key and api_key.strip() == model["default_key"]:
            session_class = model["session_class"]
            logger.info("[AI] 选择模型: %s (用户自带 key)", model["name"])
            break

    # fallback
    if session_class is None:
        model = AI_MODELS["DeepSeek"]
        session_class = model["session_class"]
        api_key = model["default_key"]
        logger.info("[AI] 使用默认 DeepSeek key")
        # 回写
        df.loc[df["userId"] == user_id, "api_key"] = api_key
        df.to_excel(USER_FILE, index=False, engine="openpyxl")

    # 创建会话并缓存
    session_obj = session_class(api_key=api_key)
    CURRENT_AI_SESSION[user_id] = (session_obj, api_key)  # ✅ 只给字典的 key 赋值

    return session_obj, api_key


def read_past_feature(path: str, index: int) -> dict[str, float]:
    """读取过去某文件的MF/MNF/RMS特征值（指定最后一行）"""
    if not os.path.exists(path):
        return {"MF": 1.0, "MNF": 1.0, "RMS": 1.0}

    df = pd.read_excel(path, engine="openpyxl")

    def get_value_safe(col_name: str) -> float:
        try:
            return float(df[col_name].iloc[-1])
        except Exception:
            return 1.0

    return {
        "MF":  get_value_safe(f"{index}_MF"),
        "MNF": get_value_safe(f"{index}_MNF"),
        "RMS": get_value_safe(f"{index}_RMS")
    }


def build_std_training(user_dir: str) -> tuple[list[dict[str, any]], str]:
    """
    生成 std_training_data 列表，并计算过去1/3/7天的 RMS 比值作为训练进度
    """
    assess_dir = os.path.join(user_dir, "Mulscle_assessment")

    today_path = latest_file(os.path.join(assess_dir, "*.xlsx"))
    logger.info("读取标准评估表: %s", today_path)
    today_df = pd.read_excel(today_path, engine="openpyxl")

    today = datetime.datetime.today()
    def date_str(d: datetime.datetime) -> str:
        return d.strftime("%Y年%m月%d日") + ".xlsx"

    def get_ratio(index: int, days_ago: int) -> float:
        past_date = today - datetime.timedelta(days=days_ago)
        past_path = os.path.join(assess_dir, date_str(past_date))
        past_features = read_past_feature(past_path, index)
        try:
            return float(today_df[f"{index}_RMS"].iloc[-1]) / past_features["RMS"]
        except Exception:
            return 1.0

    std_records: list[dict[str, Any]] = []
    for i in (1, 2):
        std_records.append({
            "MF": float(today_df[f"{i}_MF"].iloc[-1]),
            "MNF": float(today_df[f"{i}_MNF"].iloc[-1]),
            "RMS": float(today_df[f"{i}_RMS"].iloc[-1]),
            "yesterday_progress":      get_ratio(i, 1),
            "three_days_ago_progress": get_ratio(i, 3),
            "seven_days_ago_progress": get_ratio(i, 7),
        })
        logger.debug("std_record #%d: %s", i, std_records[-1])

    return std_records, today_path


@app.route("/api/generate-report", methods=["POST"])
def generate_report():
    """
    前端 POST {"id": 123} 触发生成报告
    """
    user_id = request.json.get("id")
    if user_id is None:
        return jsonify(error="缺少参数 id"), 400

    logger.info("=== 生成报告流程开始: id=%s ===", user_id)

    # 1) 基本信息
    try:
        user_row = read_user_row(user_id)
    except Exception as err:
        logger.exception("读取用户信息失败")
        return jsonify(error=str(err)), 400

    name   = user_row["name"]
    age    = int(user_row["age"])
    gender = user_row["gender"]
    ache   = user_row["ache"]
    user_dir = os.path.join(STD_DIR, name)
    logger.info("用户目录: %s", user_dir)

    # 2) rehab_training_data
    try:
        rehab_data = build_rehab_training(user_dir, user_row)
    except Exception as err:
        logger.exception("康复训练数据读取失败")
        return jsonify(error=f"训练数据读取失败: {err}"), 500

    # 3) std_training_data
    try:
        std_data, assess_path = build_std_training(user_dir)
    except Exception as err:
        logger.exception("标准评估数据读取失败")
        return jsonify(error=f"标准评估数据读取失败: {err}"), 500

    # 4) AI 会话
    try:
        ds, api_key = get_ai_session(user_id)
        logger.info("[AI] 会话已准备")
    except Exception as err:
        logger.exception("AI 会话初始化失败")
        return jsonify(error=f"AI 会话初始化失败: {err}"), 500

    # 5) 发送给 AI
    payload = {
        "all_amount": {
            "patient_name":        name,
            "age":                 age,
            "gender":              gender,
            "ache":                ache,
            "rehab_training_data": rehab_data,
            "std_training_data":   std_data
        }
    }
    
    logger.info("payload: %s", json.dumps(payload, indent=4))  # 美化输出
    logger.info("发送给 AI 的 payload 构建完成")

    try:
        report_json = ds.send_input(payload)
        logger.info("AI 返回成功")
    except Exception as err:
        logger.exception("AI 调用失败")
        return jsonify(error=f"AI 调用失败: {err}"), 500
    
    user_name = get_user_name(user_id)

    # 6) 生成最终报告
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    date_str_1 = datetime.datetime.now().strftime("%Y年%m月%d日")
    excel_path = os.path.join(user_dir, "training", "Training_data", f"{date_str}_Training-data.xlsx")
    std_paths  = [
        os.path.join(user_dir, "Mulscle_assessment", "Training_data",
                     f"{date_str_1}_Training-data_1.xlsx"),
        os.path.join(user_dir, "Mulscle_assessment", "Training_data",
                     f"{date_str_1}_Training-data_2.xlsx"),
    ]

    try:
        logger.info("调用 generate_all_from_excel")
        generate_all_from_excel(excel_path, std_paths, report_json, user_name)
        logger.info("报告生成成功")
    except Exception as err:
        logger.exception("报告生成失败")
        return jsonify(error=f"报告生成失败: {err}"), 500

    logger.info("=== 报告流程完成: id=%s ===", user_id)
    return jsonify(
        message     = "报告已生成"
    ), 200
    

@app.route('/api/get_latest_report', methods=['POST'])
def get_latest_report():
    """
    根据用户的 id 查找对应 Reports 文件夹中的最新文件
    接收参数：
    {
        "id": "用户的id"
    }
    返回：
    {
        "file_name": "最新文件的名称",
        "file_url": "文件下载的 URL"
    }
    """
    try:
        data = request.get_json()
        user_id = data.get("id")
        
        # Log input data
        logger.info(f"Received request to get latest report for user_id: {user_id}")

        if not user_id:
            logger.error("Missing 'id' parameter in request")
            return jsonify({"error": "缺少id参数"}), 400
        
        user_name = get_user_name(user_id)  # 获取用户名
        logger.info(f"Retrieved user name: {user_name} for user_id: {user_id}")

        base_dir = os.path.abspath(os.path.dirname(__file__))
        user_folder = os.path.join(base_dir, "user_files", user_name, "Reports")
        logger.info(f"Checking if user folder exists: {user_folder}")

        # 检查用户的 Reports 文件夹是否存在
        if not os.path.exists(user_folder):
            logger.error(f"Reports folder for user {user_id} does not exist: {user_folder}")
            return jsonify({"error": f"用户 {user_id} 的 Reports 文件夹不存在"}), 404

        # 获取 Reports 文件夹中的所有 PDF 文件
        files = glob.glob(os.path.join(user_folder, "*.pdf"))
        logger.info(f"Found {len(files)} PDF files in {user_folder}")

        if not files:
            logger.error(f"No PDF files found in Reports folder for user {user_id}")
            return jsonify({"error": "该文件夹下没有 PDF 文件"}), 404

        # 获取最新修改的文件
        latest_file = max(files, key=os.path.getmtime)
        logger.info(f"Latest modified file: {latest_file}")

        # 获取文件名（不带路径）
        file_name = os.path.basename(latest_file)
        logger.info(f"Extracted file name: {file_name}")

        # 构造文件的下载 URL
        file_url = f"http://115.190.118.22:5000/api/download/{user_name}/{file_name}"
        logger.info(f"Generated file URL: {file_url}")

        # 返回文件的名称和下载链接
        return jsonify({
            "file_name": file_name,
            "file_url": file_url
        }), 200

    except Exception as e:
        logger.exception(f"Error occurred while generating report for user {user_id}")
        return jsonify({'success': False, 'message': str(e)}), 500


#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#医生端
# 获取患者数据的接口
@app.route('/api/get_patient_data', methods=['GET'])
def get_patient_data():
    try:
        # 获取前端传入的用户 ID
        user_id = request.args.get('doctorId')
        
        if user_id is None:
            return jsonify({"error": "User ID is required"}), 400

        # 获取当前路径
        CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
        
        # 创建 doc_id 文件夹路径
        doc_id_folder = os.path.join(CURRENT_DIR, 'user_files', str(user_id))
        
        # 如果文件夹不存在，创建文件夹
        if not os.path.exists(doc_id_folder):
            os.makedirs(doc_id_folder)
            print(f"文件夹 {doc_id_folder} 已创建。")

        # 构造 patient_list.json 路径
        patient_list_path = os.path.join(doc_id_folder, 'patient_list.json')

        # 如果文件存在，读取文件
        if os.path.exists(patient_list_path):
            with open(patient_list_path, 'r', encoding='utf-8') as file:
                patient_list = json.load(file)
                
                # 提取患者的ID并获取患者名字
                patient_data_with_names = []
                for patient_id in patient_list:
                    user_name = get_user_name(patient_id)
                    patient_data_with_names.append({
                        "patient_id": patient_id,
                        "patient_name": user_name
                    })
                
                # 返回带有患者名字的患者数据
                return jsonify({"patient_data": patient_data_with_names})
        else:
            # 如果文件不存在，返回错误信息
            return jsonify({"error": f"文件 {patient_list_path} 不存在"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500



#获取用户的pdf列表
@app.route('/api/reports', methods=['GET'])
def get_reports_for_patient():
    # 获取前端传来的用户 ID
    user_id = request.args.get('patientId')
    
    if user_id is None:
        return jsonify({'success': False, 'message': '用户 ID 是必需的'}), 400

    # 获取用户文件夹路径
    user_name = get_user_name(user_id)
    reports_dir = os.path.join('user_files', user_name, 'Reports')

    # 检查文件夹是否存在
    if not os.path.exists(reports_dir):
        return jsonify({'success': False, 'message': f'用户 {user_name} 的报告文件夹不存在'}), 404

    try:
        # 获取 Reports 文件夹下的所有 PDF 文件
        pdf_list = [f for f in os.listdir(reports_dir) if f.endswith('.pdf')]
        
        # 如果没有 PDF 文件，返回空列表
        if not pdf_list:
            return jsonify({'success': True, 'message': '没有找到报告文件', 'reports': []})
        
        return jsonify({'success': True, 'reports': pdf_list})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/download/<user_name>/<filename>', methods=['GET'])
def download_pdf(user_name, filename):
    try:
        # 获取用户文件夹路径
        reports_dir = os.path.join('user_files', user_name, 'Reports')  # 假设每个患者有一个 Reports 文件夹存放报告
        
        # 打印调试信息，检查路径
        print(f"Trying to download from path: {reports_dir}")
        
        # 构建文件路径
        file_path = os.path.join(reports_dir, filename)
        
        # 打印文件路径检查
        print(f"File path: {file_path}")

        # 检查文件是否存在
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        # 提供文件下载
        return send_from_directory(reports_dir, filename, as_attachment=True)
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/showadvice/<user_name>', methods=['GET'])
def get_advice(user_name):
    # 构建 data.xlsx 路径
    file_path = os.path.join('user_files', user_name, 'Reports', 'data.xlsx')

    # 若文件不存在，返回空列表
    if not os.path.exists(file_path):
        return jsonify({'success': True, 'advice': []})

    try:
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data = []

        # 遍历数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row = [日期, 报告名称, 医生评价, 医生名字, 医生态度]
            if len(row) >= 5:
                formatted_date = row[0].strftime('%Y-%m-%d %H:%M') if hasattr(row[0], 'strftime') else str(row[0])
                data.append({
                    'date': formatted_date,
                    'report_name': row[1],
                    'doctor_feedback': row[2],
                    'doctor_name': row[3],
                    'doc_judge': row[4]
                })

        # ✅ 按时间倒序排序（最新在最前面）
        data = sorted(data, key=lambda x: x['date'], reverse=True)

        return jsonify({'success': True, 'advice': data})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/advice/<user_name>', methods=['POST'])
def add_advice(user_name):
    import os
    import datetime
    data = request.get_json()
    report_name = data.get('report_name')  # PDF 文件名和报告名称是相同的
    advice = data.get('advice')
    doctor_name = data.get('doctor_name')  # 医生名字字段
    doc_judge = data.get('doc_judge')      # 医生评分字段

    # 字段完整性校验
    if not report_name or not advice or not doctor_name or doc_judge is None:
        return jsonify({'success': False, 'message': '报告名、建议内容、医生名字或医生评分不能为空'}), 400

    # 找出最新报告文件夹路径
    patient_folder = os.path.join('user_files', user_name, 'Reports')
    if not os.path.isdir(patient_folder):
        return jsonify({'success': False, 'message': '患者报告文件夹不存在'}), 404

    # 获取所有报告文件，排除 data.xlsx
    all_files = sorted([f for f in os.listdir(patient_folder)
                        if os.path.isfile(os.path.join(patient_folder, f)) and f.lower() != 'data.xlsx'])

    if not all_files:
        return jsonify({'success': False, 'message': '没有找到报告文件'}), 404

    # 最新报告文件
    latest_report = all_files[-1]

    # 写入 user_files/{user_name}/Reports/data.xlsx
    file_path = os.path.join(patient_folder, 'data.xlsx')

    try:
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['日期', '报告名称', '医生评价', '医生名字', '医生评分'])
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        ws.append([now_str, report_name, advice, doctor_name, doc_judge])
        wb.save(file_path)

        # 如果是最新报告，更新根目录下的 User.xlsx
        if report_name == latest_report:
            user_xlsx = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'User.xlsx')
            if os.path.exists(user_xlsx):
                wb2 = openpyxl.load_workbook(user_xlsx)
                ws2 = wb2.active

                headers = [cell.value for cell in ws2[1]]
                header_map = {key: idx + 1 for idx, key in enumerate(headers)}

                # 若缺失字段则补上
                required_fields = ['name', 'doc_com', 'pdf', 'doc_judge']
                for field in required_fields:
                    if field not in header_map:
                        ws2.cell(row=1, column=len(headers) + 1).value = field
                        header_map[field] = len(headers) + 1
                        headers.append(field)

                updated = False
                for row in ws2.iter_rows(min_row=2):
                    if row[header_map['name'] - 1].value == user_name:
                        stored_pdf = row[header_map['pdf'] - 1].value
                        stored_pdf_time = datetime.datetime.strptime(stored_pdf.split("_")[0], '%Y%m%d%H%M%S') if stored_pdf else None
                        current_pdf_time = datetime.datetime.strptime(report_name.split("_")[0], '%Y%m%d%H%M%S')

                        if not stored_pdf or stored_pdf_time < current_pdf_time:
                            row[header_map['doc_com'] - 1].value = advice
                            row[header_map['pdf'] - 1].value = report_name
                            row[header_map['doc_judge'] - 1].value = doc_judge
                        elif stored_pdf_time == current_pdf_time:
                            prev = row[header_map['doc_com'] - 1].value or ''
                            row[header_map['doc_com'] - 1].value = (prev + '\n' + advice).strip()
                            row[header_map['doc_judge'] - 1].value = doc_judge
                        updated = True
                        break

                # 如果没有匹配用户则追加一行
                if not updated:
                    new_row = [''] * len(headers)
                    new_row[header_map['name'] - 1] = user_name
                    new_row[header_map['doc_com'] - 1] = advice
                    new_row[header_map['pdf'] - 1] = report_name
                    new_row[header_map['doc_judge'] - 1] = doc_judge
                    ws2.append(new_row)

                wb2.save(user_xlsx)

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500



#//////////////////////////////////////////////////////////////////////////////////////////////////
@app.route('/api/history/<id>', methods=['GET'])
def get_history(id):
    # 1. 构建 Reports 目录路径
    user_name = get_user_name(id)
    reports_dir = os.path.join('user_files', user_name, 'Reports')
    
    if not os.path.isdir(reports_dir):
        return jsonify({'success': True, 'history': []})

    # 2. 列出所有报告文件（排除 data.xlsx）
    all_files = sorted([
        f for f in os.listdir(reports_dir)
        if os.path.isfile(os.path.join(reports_dir, f)) and f.lower() != 'data.xlsx'
    ], key=lambda x: os.path.getctime(os.path.join(reports_dir, x)))  # 按时间正序
    

    # 3. 读取 data.xlsx，把评价按 report_name 和医生名字 分组
    advice_map = {}
    data_xlsx = os.path.join(reports_dir, 'data.xlsx')
    
    if os.path.exists(data_xlsx):
        wb = openpyxl.load_workbook(data_xlsx)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row = [日期, 报告名称, 医生评价, 医生名字, 医生态度]
            if len(row) >= 5:
                rpt, fb, doctor_name, doc_judge = row[1], row[2], row[3], row[4]
                date_str = row[0].strftime('%Y-%m-%d %H:%M') if hasattr(row[0], 'strftime') else str(row[0])

                advice_map.setdefault(rpt, []).append({
                    'date': date_str,
                    'doctor_feedback': fb,
                    'doctor_name': doctor_name,
                    'doc_judge': doc_judge
                })

    # 4. 组装返回列表
    history = []
    for rpt in all_files:
        history.append({
            'report_name': rpt,
            'download_url': f'/api/download/{user_name}/{rpt}',
            'advice': advice_map.get(rpt, [])
        })

    return jsonify({'success': True, 'history': history})


#/////////////////////////////////////////////////////////////////////
#医生简历上传
@app.route('/api/save_doctor', methods=['POST'])
def save_doctor():
    # 获取前端传来的 JSON 数据
    data = request.get_json()
    doctor_id = data.get('id')
    name = data.get('name')
    major = data.get('major')
    school = data.get('school')
    hospital = data.get('hospital')  # 新增字段：医院
    description = data.get('description')
    detailedDescription = data.get('detailedDescription')

    if not all([doctor_id, name, major, school, hospital ,description , detailedDescription]):
        return jsonify({"error": "缺少必要字段"}), 400

    file_path = 'Doctor.xlsx'

    # 如果文件不存在，创建新文件和表头
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "name", "major", "school", "hospital" ,"description" , "detailedDescription"])
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    # 查找是否存在该 doctor_id
    updated = False
    for row in ws.iter_rows(min_row=2):  # 从第2行开始，跳过表头
        if str(row[0].value) == str(doctor_id):
            row[1].value = name
            row[2].value = major
            row[3].value = school
            row[4].value = hospital
            row[5].value = description
            row[6].value = detailedDescription
            updated = True
            break

    # 如果没有找到该ID，追加一行新数据
    if not updated:
        ws.append([doctor_id, name, major, school, hospital , description , detailedDescription])

    wb.save(file_path)

    msg = "医生信息已更新" if updated else "医生信息已新增"
    return jsonify({"message": msg}), 200

#////////////////////////////////////////////////////////////////////////
#医生添加患者
@app.route('/api/search_patients', methods=['GET'])
def search_patients():
    keyword = request.args.get('keyword', '').strip()

    if not keyword:
        return jsonify({"patients": []})

    user_file = os.path.join(os.path.dirname(__file__), 'User.xlsx')
    if not os.path.exists(user_file):
        return jsonify({"error": f"用户数据文件不存在: {user_file}"}), 500

    try:
        wb = openpyxl.load_workbook(user_file)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"无法读取Excel: {str(e)}"}), 500

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue  # 防止空行或列不足
        patient_id, name = row[0], row[1]
        if name and keyword in str(name):
            results.append({"id": patient_id, "name": name})

    return jsonify({"patients": results})

@app.route('/api/manage_patient', methods=['POST'])
def manage_patient():
    """
    JSON 请求体应包含：
    {
        "doctor_id": "D202501",
        "patient_id": "xxx",
        "action": "add" 或 "remove"
    }
    """
    data = request.get_json()
    doctor_id = data.get('doctor_id')
    patient_id = data.get('patient_id')
    action = data.get('action')

    if not doctor_id or not patient_id or action not in ('add', 'remove'):
        return jsonify({"error": "缺少参数或 action 不合法"}), 400

    # 确定路径
    doctor_folder = os.path.join('user_files', str(doctor_id))
    os.makedirs(doctor_folder, exist_ok=True)
    patient_list_path = os.path.join(doctor_folder, 'patient_list.json')

    # 加载已有列表，若不存在则创建新的空列表
    if os.path.exists(patient_list_path):
        try:
            with open(patient_list_path, 'r', encoding='utf-8') as f:
                patient_list = json.load(f)
        except (IOError, json.JSONDecodeError) as e:
            return jsonify({"error": f"读取患者列表失败: {e}"}), 500
    else:
        patient_list = []

    # 更新 User.xlsx 中的 doctor 列
    user_xlsx_path = os.path.join('User.xlsx')
    if not os.path.exists(user_xlsx_path):
        return jsonify({"error": "User.xlsx 文件不存在"}), 500

    try:
        wb_user = openpyxl.load_workbook(user_xlsx_path)
        ws_user = wb_user.active
    except Exception as e:
        return jsonify({"error": f"打开 User.xlsx 文件失败: {e}"}), 500

    # 获取对应患者的行
    patient_row = None
    for row in ws_user.iter_rows(min_row=2, values_only=False):
        if row[0].value == patient_id:  # 假设第一列是患者ID
            patient_row = row
            break

    if not patient_row:
        return jsonify({"error": "未找到该患者"}), 404

    # 获取医生名字，从 Doctor.xlsx 中查找
    doctor_name = get_doctor_name(doctor_id)  # 通过 get_doctor_name 获取医生名字

    # 根据 action 执行相应操作
    if action == 'add':
        # 查找是否已有 doctor 列，如果没有则创建
        doctor_column = None
        for col_idx, cell in enumerate(ws_user[1]):
            if cell.value == 'doctor':
                doctor_column = col_idx
                break

        if doctor_column is None:
            # 如果没有 doctor 列，则创建
            doctor_column = len(ws_user[1]) + 1
            ws_user.cell(row=1, column=doctor_column, value="doctor")

        # 检查 doctor_column 是否有效
        if doctor_column is None:
            return jsonify({"error": "User.xlsx 文件中没有 doctor 列"}), 500

        # 获取当前 doctor 列的值
        current_doctor_value = patient_row[doctor_column].value

        if not current_doctor_value:
            # 如果没有医生名，则填入医生名字
            patient_row[doctor_column].value = doctor_name
        else:
            # 如果已有医生名，则检查是否已经存在该医生名字
            doctor_names = current_doctor_value.split(", ")
            # 过滤掉 None 或空值，确保所有元素都是字符串类型
            doctor_names = [str(doctor) if doctor is not None else "" for doctor in doctor_names]
            if doctor_name not in doctor_names:
                doctor_names.append(doctor_name)
                patient_row[doctor_column].value = ", ".join(doctor_names)
            else:
                return jsonify({"message": "该医生已经存在", "doctor_name": doctor_name}), 200

    elif action == 'remove':
        # 查找 doctor 列
        doctor_column = None
        for col_idx, cell in enumerate(ws_user[1]):
            if cell.value == 'doctor':
                doctor_column = col_idx
                break

        if doctor_column is None:
            return jsonify({"error": "User.xlsx 文件中没有 doctor 列"}), 500

        # 获取当前 doctor 列的值
        current_doctor_value = patient_row[doctor_column].value

        if current_doctor_value:
            # 如果有医生名字，删除该医生的名字
            doctor_names = current_doctor_value.split(", ")
            # 过滤掉 None 或空值，确保所有元素都是字符串类型
            doctor_names = [str(doctor) if doctor is not None else "" for doctor in doctor_names]
            if doctor_name in doctor_names:
                doctor_names.remove(doctor_name)
                if doctor_names:
                    patient_row[doctor_column].value = ", ".join(doctor_names)
                else:
                    patient_row[doctor_column].value = None
            else:
                return jsonify({"error": "医生名字不存在于患者记录中"}), 400

    # 保存更新后的 User.xlsx 文件
    try:
        wb_user.save(user_xlsx_path)
    except IOError as e:
        return jsonify({"error": f"保存 User.xlsx 文件失败: {e}"}), 500

    # 执行添加或删除操作
    if action == 'add':
        if patient_id not in patient_list:
            patient_list.append(patient_id)
            try:
                with open(patient_list_path, 'w', encoding='utf-8') as f:
                    json.dump(patient_list, f, indent=2)
            except IOError as e:
                return jsonify({"error": f"保存患者列表失败: {e}"}), 500
            return jsonify({"message": "患者已添加", "patient_id": patient_id}), 200
        else:
            return jsonify({"message": "该患者已存在列表中", "patient_id": patient_id}), 200

    elif action == 'remove':
        if patient_id in patient_list:
            patient_list.remove(patient_id)
            try:
                with open(patient_list_path, 'w', encoding='utf-8') as f:
                    json.dump(patient_list, f, indent=2)
            except IOError as e:
                return jsonify({"error": f"保存患者列表失败: {e}"}), 500
            return jsonify({"message": "患者已移除", "patient_id": patient_id}), 200
        else:
            return jsonify({"message": "列表中不存在该患者", "patient_id": patient_id}), 404



# 获取医生名字的函数，从 Doctor.xlsx 中根据 doctor_id 获取
def get_doctor_name(doctor_id):
    doctor_xlsx_path = os.path.join('Doctor.xlsx')
    if not os.path.exists(doctor_xlsx_path):
        raise FileNotFoundError("Doctor.xlsx 文件未找到")

    wb_doc = openpyxl.load_workbook(doctor_xlsx_path)
    ws_doc = wb_doc.active

    # 查找医生ID对应的名字
    for row in ws_doc.iter_rows(min_row=2, values_only=True):
        if row[0] == doctor_id:  # 假设第一列是医生ID
            return row[1]  # 假设第二列是医生的名字
    
    return None  # 如果没有找到医生，返回None




if __name__ == '__main__':
    start_udp_server()
    # 确保存在必要的目录
    os.makedirs('data', exist_ok=True)
    os.makedirs('model', exist_ok=True)
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    create_excel_file()  # 确保文件存在

    app.run(debug=True, host='0.0.0.0', port=5000)
